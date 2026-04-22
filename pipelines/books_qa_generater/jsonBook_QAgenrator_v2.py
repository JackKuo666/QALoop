#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图书SFT问答对生成器 v2.0（参考demo_thesis_QA_genrator_v1_2.py）
核心流程：读取图书 → 拆分章节 → 生成SFT问答对 → 质量评估 → 保存
输出格式适用于大模型SFT训练
支持两阶段推理链生成

新增功能：
- 自动质量评分系统（QualityScorer）
- SimHash去重功能
- 质量过滤功能
- 图书/Chunk/QA多级统计
- 详细的统计信息输出
"""
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import json
import logging
import os
import re
import sys
import time
import uuid
from typing import List, Dict, Any, Optional, Tuple, Union
from openai import OpenAI
from dotenv import load_dotenv
import argparse
import hashlib
from collections import defaultdict
import threading

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        import openpyxl
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

# 线程状态管理器
class ThreadStatusManager:
    def __init__(self, max_threads: int = 1):
        self.max_threads = max_threads
        self.status = {}  # {thread_id: {'file': str, 'qa_count': int, 'status': str}}
        self.lock = threading.Lock()
        self.running = True

    def update_status(self, thread_id: int, file: str, qa_count: int, status: str = "处理中"):
        with self.lock:
            self.status[thread_id] = {
                'file': file,
                'qa_count': qa_count,
                'status': status
            }

    def get_status(self) -> Dict[int, Dict]:
        with self.lock:
            return self.status.copy()

    def stop(self):
        self.running = False

# 全局状态管理器
status_manager = ThreadStatusManager(max_threads=1)

# 添加python目录到sys.path，以便导入reasoning_diversity模块
_script_dir = os.path.dirname(os.path.abspath(__file__))
_python_dir = os.path.join(os.path.dirname(_script_dir), 'python')
if _python_dir not in sys.path:
    sys.path.insert(0, _python_dir)

try:
    from reasoning_diversity import diversity_filter_qas
except ImportError:
    diversity_filter_qas = None

def assign_curriculum_stage(difficulty: str, question_cot: str) -> int:
    """
    stage 1: easy + 短推理
    stage 2: medium + 中等推理
    stage 3: hard 或 长推理
    """
    difficulty = (difficulty or "medium").lower().strip()
    steps = 0
    if question_cot:
        steps = len([s for s in question_cot.split("\n") if s.strip()])

    if difficulty == "easy" and steps <= 3:
        return 1
    if difficulty in ["easy", "medium"] and 3 < steps <= 5:
        return 2
    return 3




# 加载 .env 文件（优先脚本目录）
script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, '.env')
load_dotenv(env_path)

# --- 配置 ---
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
DEFAULT_MODEL = os.getenv("DEFAULT_MODEL", "gpt-5.1")

# --- 常量配置 ---
MIN_CHAPTER_LENGTH = 100
MAX_CHUNK_SIZE = 200
IDEA_CHUNK_LENGTH = 2500
MAX_Q_PER_CHUNK = 40
BOOK_SOURCES = {'book', 'textbook', 'manual', 'guide', 'encyclopedia', 'reference'}
OVER_GENERATE_FACTOR = 2  # 超生成因子
REASONING_SECTIONS = ['results', 'discussion', 'methods', 'background', 'introduction']  # 适合推理链的章节

# --- 结束标记关键词（遇到这些关键词时停止提取内容）---
END_SECTION_KEYWORDS = [
    '参考文献', '参考书目', 'References', 'REFERENCE', 'REFERENCES',
    'Bibliography', 'BIBLIOGRAPHY', '附录', 'Appendix', 'APPENDIX',
    '后记', '致谢', 'Acknowledgments', 'ACKNOWLEDGMENTS', '索引', 'Index', 'INDEX'
]

# --- 违禁短语配置 ---
FORBIDDEN_PHRASES = [
    "文中指出", "文中提到", "文中认为", "文中表明",
    "本文指出", "本文认为", "本文表明", "本文中", "文本中", "实验中", "试验中",
    "文章指出", "文章认为", "文章表明", "文章中",
    "本研究指出", "本研究认为", "本研究表明", "本研究", "研究中",
    "该研究指出", "该研究认为", "该研究表明", "该研究", "该章节",
    "在该实验中", "在本实验中", "在本研究中", "在这项研究中",
    "在这篇文章中", "在该文章中", "在该论文中", "在上述研究中",
    "根据给定内容", "根据上述内容", "根据文中内容", "根据本文内容", "根据文本内容",
    "作者认为", "作者指出", "作者提到",
    "根据给出的", "根据给出的文本", "根据给出的内容",
    "这段文字", "这段文本", "这段内容", "这段研究", "这段描述",
]

# --- 日志配置 ---
def setup_logger():
    log_dir = os.path.join(os.getcwd(), 'log')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'gen_book_sft_qa.log')

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    if logger.hasHandlers():
        logger.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    console_handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger

logger = setup_logger()

# --- 线程本地存储 ---
_thread_local = threading.local()

def get_thread_id():
    """获取当前线程ID"""
    return getattr(_thread_local, 'thread_id', 0)

def set_thread_id(thread_id):
    """设置当前线程ID"""
    _thread_local.thread_id = thread_id

def tprint(message):
    """带线程ID的打印"""
    thread_id = get_thread_id()
    prefix = f"[T{thread_id}]" if thread_id else ""
    print(f"{prefix} {message}", flush=True)

# --- LLM调用统计 ---
_llm_call_counter = 0
_llm_call_lock = threading.Lock()

def increment_llm_call_count():
    """增加LLM调用计数（线程安全）"""
    global _llm_call_counter
    with _llm_call_lock:
        _llm_call_counter += 1
        return _llm_call_counter

def get_llm_call_count():
    """获取LLM调用总次数"""
    global _llm_call_counter
    with _llm_call_lock:
        return _llm_call_counter

def reset_llm_call_count():
    """重置LLM调用计数"""
    global _llm_call_counter
    with _llm_call_lock:
        _llm_call_counter = 0

# --- 工具函数 ---
def clean_text_basic(text: str) -> str:
    """基础文本清理"""
    if not text:
        return ""
    # 删除图片
    text = re.sub(r'!\[[^\]]*?\]\([^\)]*?\)', ' ', text)
    # 删除HTML标签（保留sub和sup）
    text = re.sub(r'<(?!sub|sup|/sub|/sup)[^>]+>', ' ', text)
    # 删除markdown链接
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    # 删除表格
    text = re.sub(r'^\s*\|.*\|\s*$', ' ', text, flags=re.MULTILINE)
    # 合并空行
    text = re.sub(r'\n{3,}', '\n\n', text)
    # 合并空格
    text = re.sub(r'[ \t]{2,}', ' ', text)
    return text.strip()

def sanitize_text_forbidden_phrases(text: str) -> str:
    """清洗违禁短语"""
    if not text:
        return text
    text = re.sub(r'\n+', ' ', text)
    text = re.sub(r'\s{2,}', ' ', text)
    text = re.sub(r'[，。；：]{2,}', lambda m: m.group(0)[0], text)
    text = re.sub(r'^[，；：、\s]+', '', text)
    return text.strip()

def is_study_dependent(text: str) -> bool:
    """判断文本是否依赖具体研究"""
    if not text:
        return False
    patterns = [
        r"根据(本|该|这项)?研究",
        r"在(本|该|这项)?研究中",
        r"(本|该)研究表明",
        r"(本文|该文)(中)?(认为|指出|表明|描述)",
        r"(结果|摘要|讨论|方法)(部分)?指出",
        r"在(本|该)?实验中",
        r"(本|该)实验(中)?",
        r"该论文(中|指出|认为|表明)",
    ]
    for pattern in patterns:
        if re.search(pattern, text):
            return True
    return False

def estimate_tokens(text: str) -> int:
    """估算token数"""
    if not text:
        return 0
    words = len(text.split())
    return int(words * 1.3)

def calculate_qa_count_per_chunk(length: int, min_q: int = 3, max_q: int = 40, base_length: int = 2000) -> int:
    """
    根据length字段自动计算应该生成的QA数量
    规则：
    - length < 100：不生成QA（返回0）
    - 100-200：生成1个QA
    - 200-2000：生成2个QA
    - 2000以上：按照 int(length / 2000) + 1 计算QA数量

    Args:
        length: 章节长度（从JSON的length字段读取）
        min_q: 最小QA数量（默认3，已不使用）
        max_q: 最大QA数量（默认40）
        base_length: 基准文本长度，每base_length字符对应1个QA（默认2000）

    Returns:
        int: 应该生成的QA数量
    """
    if not length or length <= 0:
        return 0

    # 根据length分段计算QA数量
    if length < 100:
        # length < 100：不生成QA
        return 0
    elif length < 200:
        # 100-200：生成1个QA
        return 1
    elif length < 2000:
        # 200-2000：生成2个QA
        return 2
    else:
        # 2000以上：按照 int(length / 2000) + 1 计算QA数量
        target_qa = int(length / base_length) + 1
        # 应用最大QA数量限制
        target_qa = min(max_q, target_qa)
        return target_qa

def contains_forbidden_phrases(text: str) -> bool:
    """检查是否包含违禁短语"""
    if not text:
        return False
    return any(phrase in text for phrase in FORBIDDEN_PHRASES)

# ==============================================================================
# SimHash 去重（轻量、无依赖）
# ==============================================================================
def _tokenize_for_simhash(text: str) -> List[str]:
    """
    中文不适合 split()；这里用'汉字2-gram + 英文词'混合。
    """
    if not text:
        return []
    text = re.sub(r'\s+', ' ', text).strip()
    en_words = re.findall(r"[A-Za-z0-9]+", text.lower())
    zh_chars = re.findall(r"[\u4e00-\u9fff]", text)
    zh_ngrams = ["".join(zh_chars[i:i+2]) for i in range(max(0, len(zh_chars)-1))]
    return en_words + zh_ngrams

def simhash64(tokens: List[str]) -> int:
    if not tokens:
        return 0
    v = [0] * 64
    for tok in tokens:
        h = int(hashlib.md5(tok.encode("utf-8")).hexdigest(), 16)
        for i in range(64):
            bit = (h >> i) & 1
            v[i] += 1 if bit else -1
    out = 0
    for i in range(64):
        if v[i] > 0:
            out |= (1 << i)
    return out

def hamming64(a: int, b: int) -> int:
    return (a ^ b).bit_count()

def dedup_qas_simhash(qas: List[Dict[str, Any]], max_hamming: int = 6) -> List[Dict[str, Any]]:
    """
    基于 question 的 SimHash 去重（阈值越小越严格）。
    """
    if not qas:
        return []

    kept = []
    hashes: List[int] = []
    for qa in qas:
        q = qa.get("question", "") or ""
        h = simhash64(_tokenize_for_simhash(q))
        dup = False
        for prev in hashes:
            if hamming64(h, prev) <= max_hamming:
                dup = True
                break
        if not dup:
            kept.append(qa)
            hashes.append(h)
    return kept

# ==============================================================================
# 自动质量评分系统
# ==============================================================================
class QualityScorer:
    def __init__(self):
        self.quality_thresholds = {
            'excellent': 90,
            'good': 75,
            'acceptable': 60,
            'poor': 40
        }

        self.quality_penalty_patterns = {
            'vague_reference': [
                r'文中', r'本文', r'文章中', r'该章节', r'该部分',
                r'本图书', r'本书', r'我们发现', r'我们得出'
            ],
            'incomplete': [
                r'等等', r'其他', r'各种', r'诸多', r'若干',
                r'相关', r'有关', r'一定', r'某些', r'部分'
            ],
            'redundant': [
                r'非常重要的', r'至关重要的', r'极其重要的',
                r'非常关键的', r'极其关键的'
            ],
            'academic_violation': [
                r'根据本实验', r'通过本实验', r'本研究表明',
                r'如图所示', r'见表', r'见图'
            ]
        }

        self.score_weights = {
            'length_score': 0.10,
            'completeness_score': 0.15,
            'logic_score': 0.20,
            'clarity_score': 0.15,
            'academic_score': 0.15,
            'reasoning_score': 0.25
        }

    def score_qa_pair(self, qa: Dict[str, Any]) -> Dict[str, Any]:
        question = qa.get('question', '') or ''
        answer = qa.get('answer', '') or ''
        reasoning_steps = qa.get('reasoning_steps', []) or []
        question_cot = qa.get('question_cot', '') or ''
        final_conclusion = qa.get('final_conclusion', '') or ''

        length_score = self._score_length(question, answer)
        completeness_score = self._score_completeness(question, answer, reasoning_steps)
        logic_score = self._score_logic(answer, reasoning_steps, question_cot)
        clarity_score = self._score_clarity(question, answer)
        academic_score = self._score_academic_compliance(question, answer)
        reasoning_score = self._score_reasoning_chain(reasoning_steps, question_cot, final_conclusion)

        total_score = (
            length_score * self.score_weights['length_score'] +
            completeness_score * self.score_weights['completeness_score'] +
            logic_score * self.score_weights['logic_score'] +
            clarity_score * self.score_weights['clarity_score'] +
            academic_score * self.score_weights['academic_score'] +
            reasoning_score * self.score_weights['reasoning_score']
        )

        quality_level = self._determine_quality_level(total_score)
        return {
            'total_score': round(total_score, 2),
            'quality_level': quality_level,
            'dimension_scores': {
                'length_score': round(length_score, 2),
                'completeness_score': round(completeness_score, 2),
                'logic_score': round(logic_score, 2),
                'clarity_score': round(clarity_score, 2),
                'academic_score': round(academic_score, 2),
                'reasoning_score': round(reasoning_score, 2),
            },
            'weights': self.score_weights,
            'issues': self._identify_issues(question, answer, reasoning_steps),
            'suggestions': self._generate_suggestions(total_score, reasoning_steps),
        }

    def _score_length(self, question: str, answer: str) -> float:
        q_len, a_len = len(question), len(answer)

        # 问题：8~120 最佳
        if 8 <= q_len <= 60:
            q_score = 100
        elif 60 < q_len <= 120:
            q_score = 90
        elif 120 < q_len <= 200:
            q_score = 80
        else:
            q_score = max(0, 100 - abs(q_len - 80) * 0.5)

        # 答案：60~500 最佳
        if 60 <= a_len <= 220:
            a_score = 100
        elif 220 < a_len <= 500:
            a_score = 92
        elif 500 < a_len <= 900:
            a_score = 85
        else:
            a_score = max(0, 100 - abs(a_len - 320) * 0.06)

        return (q_score + a_score) / 2

    def _soft_relevance(self, q: str, a: str) -> float:
        """
        中文不分词时用 2-gram 近似相关度；英文用词。
        """
        qt = set(_tokenize_for_simhash(q))
        at = set(_tokenize_for_simhash(a))
        if not qt:
            return 0.0
        return len(qt & at) / max(1, len(qt))

    def _score_completeness(self, question: str, answer: str, reasoning_steps: List[str]) -> float:
        score = 100.0

        rel = self._soft_relevance(question, answer)
        if rel < 0.08:
            score -= 35
        elif rel < 0.15:
            score -= 20
        elif rel < 0.25:
            score -= 10

        # 推理链完整性
        if reasoning_steps:
            if len(reasoning_steps) < 3:
                score -= 20
            elif len(reasoning_steps) > 10:
                score -= 10
            else:
                score += 8

        # 模糊表述扣分
        for pat in [r'等等', r'其他', r'各种', r'相关', r'一定', r'可能']:
            if re.search(pat, answer):
                score -= 10

        return max(0, min(100, score))

    def _score_logic(self, answer: str, reasoning_steps: List[str], question_cot: str) -> float:
        score = 100.0

        # 答案逻辑结构：至少 2 句更好
        sents = [s.strip() for s in re.split(r'[。！？]', answer) if s.strip()]
        if len(sents) < 2:
            score -= 20
        elif len(sents) > 10:
            score -= 8

        # 推理链连贯性（轻量）
        chain_text = ""
        if reasoning_steps:
            chain_text = "\n".join(reasoning_steps)
        elif question_cot:
            chain_text = question_cot

        if chain_text:
            connectors = ['因此', '所以', '从而', '进而', '导致', '由于', '因而', '首先', '其次', '最后']
            conn_hits = sum(1 for c in connectors if c in chain_text)
            if conn_hits == 0:
                score -= 12

        return max(0, min(100, score))

    def _score_clarity(self, question: str, answer: str) -> float:
        score = 100.0

        # 超长句扣分
        for sent in [s.strip() for s in re.split(r'[。！？]', question) if s.strip()]:
            if len(sent) > 120:
                score -= 8

        for sent in [s.strip() for s in re.split(r'[。！？]', answer) if s.strip()]:
            if len(sent) > 180:
                score -= 5

        # 可疑符号扣分
        for marker in ['...', '？？', '。。', '??', '!!!']:
            if marker in question or marker in answer:
                score -= 10

        return max(0, min(100, score))

    def _score_academic_compliance(self, question: str, answer: str) -> float:
        score = 100.0

        # 违禁模式扣分
        for category, patterns in self.quality_penalty_patterns.items():
            for pattern in patterns:
                if re.search(pattern, question) or re.search(pattern, answer):
                    if category == 'academic_violation':
                        score -= 22
                    elif category == 'vague_reference':
                        score -= 14
                    else:
                        score -= 8

        # 学术指标加分
        academic_indicators = ['机制', '原理', '方法', '理论', '规律', '特征', '因素', '假设', '模型', '推导', '约束']
        academic_count = sum(1 for ind in academic_indicators if ind in answer)
        if academic_count >= 3:
            score += 8
        elif academic_count == 0:
            score -= 8

        return max(0, min(100, score))

    def _score_reasoning_chain(self, reasoning_steps: List[str], question_cot: str, final_conclusion: str) -> float:
        if not reasoning_steps and not question_cot:
            return 0.0

        score = 100.0

        # steps 数量
        if reasoning_steps:
            n = len(reasoning_steps)
            if n < 3:
                score -= 25
            elif n > 9:
                score -= 10
            else:
                score += 8

            avg_len = sum(len(s) for s in reasoning_steps) / max(1, n)
            if avg_len < 12:
                score -= 18
            elif avg_len > 100:
                score -= 8

        # question_cot 句子数量
        if question_cot:
            c_sents = [s.strip() for s in re.split(r'[。！？\n]', question_cot) if s.strip()]
            if len(c_sents) < 3:
                score -= 10
            elif len(c_sents) > 12:
                score -= 8

        # 结论
        if final_conclusion:
            if len(final_conclusion) < 10:
                score -= 8
            elif len(final_conclusion) > 140:
                score -= 4

        return max(0, min(100, score))

    def _determine_quality_level(self, score: float) -> str:
        if score >= self.quality_thresholds['excellent']:
            return '优秀'
        if score >= self.quality_thresholds['good']:
            return '良好'
        if score >= self.quality_thresholds['acceptable']:
            return '可接受'
        return '较差'

    def _identify_issues(self, question: str, answer: str, reasoning_steps: List[str]) -> List[str]:
        issues = []
        if len(question) < 8:
            issues.append('问题过短')
        elif len(question) > 220:
            issues.append('问题过长')

        if len(answer) < 30:
            issues.append('答案过短')
        elif len(answer) > 1200:
            issues.append('答案过长')

        if contains_forbidden_phrases(question) or contains_forbidden_phrases(answer):
            issues.append('存在违禁短语')

        if is_study_dependent(question) or is_study_dependent(answer):
            issues.append('存在研究依赖表达')

        if reasoning_steps:
            if len(reasoning_steps) < 3:
                issues.append('推理步骤不足')
            elif len(reasoning_steps) > 10:
                issues.append('推理步骤过多')

        return issues

    def _generate_suggestions(self, score: float, reasoning_steps: List[str]) -> List[str]:
        s = []
        if score < 60:
            s.append('建议重新生成或进行失败回收再生成')
        if score < 75:
            s.append('建议检查：是否存在模糊指代/学术违规口吻/推理链不连贯')
        if not reasoning_steps and score > 80:
            s.append('可考虑启用推理链生成以提升深度')
        if score >= 90:
            s.append('质量优秀，可直接使用')
        return s

    def filter_by_quality(self, qas: List[Dict[str, Any]], min_score: float = 60.0) -> List[Dict[str, Any]]:
        filtered = []
        for qa in qas:
            report = self.score_qa_pair(qa)
            qa2 = qa.copy()
            qa2["quality_report"] = report
            if report["total_score"] >= min_score:
                filtered.append(qa2)
        return filtered

    def get_quality_statistics(self, qas: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not qas:
            return {
                'total_count': 0, 'average_score': 0, 'max_score': 0, 'min_score': 0,
                'excellent_count': 0, 'good_count': 0, 'acceptable_count': 0, 'poor_count': 0,
                'pass_rate': 0
            }

        scores = []
        for qa in qas:
            rep = qa.get("quality_report")
            if rep and isinstance(rep, dict) and "total_score" in rep:
                scores.append(rep["total_score"])

        if not scores:
            return {
                'total_count': len(qas), 'average_score': 0, 'max_score': 0, 'min_score': 0,
                'excellent_count': 0, 'good_count': 0, 'acceptable_count': 0, 'poor_count': 0,
                'pass_rate': 0
            }

        return {
            'total_count': len(qas),
            'average_score': round(sum(scores) / len(scores), 2),
            'max_score': max(scores),
            'min_score': min(scores),
            'excellent_count': sum(1 for s in scores if s >= 90),
            'good_count': sum(1 for s in scores if 75 <= s < 90),
            'acceptable_count': sum(1 for s in scores if 60 <= s < 75),
            'poor_count': sum(1 for s in scores if s < 60),
            'pass_rate': round(sum(1 for s in scores if s >= 60) / len(scores) * 100, 2)
        }

# ==============================================================================
# 两阶段推理链生成函数（参考PaperQAGenerator_v9.2）
# ==============================================================================

def is_reasoning_suitable_section(section_name: str) -> bool:
    """
    判断section是否适合使用推理链（Chain of Thought）。

    适合推理链的section需要多步推理才能得出结论，如实验结果分析、讨论等。
    不适合的section主要是事实性陈述和总结性内容。

    Args:
        section_name: 章节名称

    Returns:
        bool: True if 适合使用推理链，False otherwise
    """
    section_lower = section_name.lower().strip()

    # 检查非推理链section（结论、摘要等）
    non_reasoning_sections = ["abstract", "conclusion", "summary"]
    for non_reasoning in non_reasoning_sections:
        if non_reasoning in section_lower:
            return False

    # 检查推理链section
    for reasoning in REASONING_SECTIONS:
        if reasoning in section_lower:
            return True

    # 默认返回True（对于未明确分类的section，默认使用推理链）
    return True


def build_chain_extraction_prompt(
    section_name: str,
    section_text: str,
    max_chains: int = 3
) -> str:
    """
    构建"从章节文本抽取推理链"的大 prompt（system+user 合并）。

    Args:
        section_name: 章节名称
        section_text: 章节文本
        max_chains: 最大推理链数量

    Returns:
        str: 完整的prompt
    """
    # 对文本进行截断，避免prompt过长导致超时
    # 推理链抽取需要处理更多内容，但过长会导致超时，所以限制在30000字符
    processed_text = (
        section_text[:30000] + "\n\n[以下内容因长度被截断]"
        if len(section_text) > 30000
        else section_text
    )

    prompt = f"""你是一位农业育种与生命科学领域的专家阅读系统，擅长从图书章节片段中总结可复用的推理链。

【你的任务】
给定一个图书章节内容，请从中抽取 1~{max_chains} 条"可用于构造多步推理问答"的推理链。
每条推理链必须满足：
- 基于文本中明确出现的事实（理论阐述、方法描述、概念解释等）
- 通过 3~7 个逻辑步骤推理得出某个客观结论
- 结论是"客观可判断对错"的（如某种关系、比较、更优方案等）
- 推理过程不依赖'本章/该章节/本节'等指代表述

【输出格式】
严格输出一个 JSON 对象：
{{
  "chains": [
    {{
      "id": "C1",
      "final_conclusion": "一句话客观结论（可直接作为答案）",
      "steps": [
        "Step 1: ...",
        "Step 2: ...",
        "Step 3: ..."
      ],
      "support_facts": [
        "从文本抽取或概括的关键事实1",
        "关键事实2"
      ],
      "potential_question_templates": [
        "围绕该结论可以提问的问题模板1",
        "问题模板2"
      ]
    }}
  ]
}}

【禁止内容】
- 不要使用"该章节/本章/本节"等指代原文的措辞
- 不要引用图表编号、外部数据库编号
- 不要生成依赖于具体样本数量、具体群体数目、具体参数的结论

【章节内容】
名称：{section_name}
内容：
\"\"\"markdown
{processed_text}
\"\"\""""
    return prompt


def build_chain_to_qa_prompt(chain_json_str: str) -> str:
    """
    构建"单条推理链 → 一道需要多步推理的问答对"的 prompt。

    Args:
        chain_json_str: 单条 chain 的 JSON 字符串

    Returns:
        str: 完整的prompt
    """
    prompt = f"""你是一位农业育种与生命科学领域的教学专家，负责把结构化推理链转化为"需要多步推理才能回答的客观问答对"，用于大模型 SFT 训练。

下面是从图书章节中抽取的一条推理链（JSON）：
```json
{chain_json_str}
```

【你的任务】
基于这条推理链，构造 1 题"需要多步推理才能回答的问答对"，输出 JSON 对象：
{{
  "question": "面向研究生/科研人员、需要理解多个事实并综合推理的问题",
  "answer": "一段简明客观答案（不包含思维链）",
  "cot": [
    "Step 1: ...",
    "Step 2: ...",
    "Step 3: ...",
    "Step 4: ..."
  ],
  "meta": {{
    "difficulty": "easy | medium | hard",
    "difficulty_score": 0.0~1.0,
    "tags": ["concept", "mechanism", "method", "result", "application", "..."]
  }}
}}

【问题设计要求】

必须需要综合多个 support_facts 和推理步骤才能得出答案，不能是抄一句话即可回答。

问题要脱离原图书章节也成立，不能包含"该章节/本章/本节"等指代。

问题聚焦通用的科学关系或机制（如：哪种方法更适合解决某类问题？什么条件下会出现某种现象？）。

**重要**：如果问题中已经明确给出了所有信息（如①、②、③等编号信息，或"已知："、"基于这些信息"等表述），答案必须提供问题中未直接给出的推理过程、逻辑依据或综合结论，而不能只是简单复述问题中的信息。答案应该基于问题中的信息进行推理，得出新的结论或提供问题中未明确说明的逻辑关系。

答案必须是客观的、唯一可判断对错的结论。

【思维链（CoT）要求】

CoT（cot 数组）用 4~7 步自然语言中文推理，逐步从事实推导到结论。

**重要**：CoT应该基于推理链的抽象逻辑，而非图书中的具体数值或细节。CoT描述的是通用的科学推理过程，适用于类似的其他研究。

例如：
- ✅ 好的CoT："激素通过调节生理指标提高抗逆性"
- ❌ 差的CoT："100mg/L脱落酸处理24小时后脯氨酸含量提高35%"

【禁止内容】

不要在 question 或 answer 中使用具体数值、浓度、时间等图书特有细节。

不要在Cot中引用具体的参数、浓度、时间等图书特有细节。

【输出要求】

严格输出一个 JSON 对象（而不是数组）。

不要添加额外解释或自然语言说明。"""
    return prompt


def _is_retryable_error(error: Exception) -> bool:
    """判断错误是否可重试"""
    error_str = str(error).lower()
    error_type = type(error).__name__

    # 可重试的错误类型
    retryable_types = ['Timeout', 'ConnectionError', 'HTTPError', 'RequestException']
    retryable_keywords = ['timeout', 'connection', 'network', '5', '503', '502', '504', '500']

    # 不可重试的错误类型
    non_retryable_keywords = ['401', '403', '400', 'invalid api key', 'authentication', 'json', 'syntax']

    # 检查是否是不可重试的错误
    if any(keyword in error_str for keyword in non_retryable_keywords):
        return False

    # 检查是否是可重试的错误
    if any(keyword in error_str for keyword in retryable_keywords):
        return True

    if error_type in retryable_types:
        return True

    # 默认情况下，网络相关错误可重试，其他错误不可重试
    return False


def _extract_json_from_text(text: str) -> Optional[Dict]:
    """从文本中尝试提取JSON信息（容错处理）"""
    if not text:
        return None

    # 策略1：尝试找到JSON对象
    json_patterns = [
        r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}',  # 简单JSON对象
        r'\[[^\[\]]*(?:\[[^\[\]]*\][^\[\]]*)*\]',  # 简单JSON数组
    ]

    for pattern in json_patterns:
        matches = re.findall(pattern, text, re.DOTALL)
        for match in matches:
            try:
                return json.loads(match)
            except Exception:
                continue

    # 策略2：尝试提取关键字段
    result = {}
    if '"chains"' in text or "'chains'" in text:
        # 尝试提取chains相关信息
        chains_match = re.search(r'chains["\']?\s*:\s*\[', text, re.IGNORECASE)
        if chains_match:
            result['has_chains'] = True

    if '"question"' in text or "'question'" in text:
        result['has_question'] = True

    if '"answer"' in text or "'answer'" in text:
        result['has_answer'] = True

    return result if result else None


def call_responses_for_json(
    prompt: str,
    model: str = DEFAULT_MODEL,
    max_output_tokens: int = 8000,
    max_retries: int = 3,
) -> Any:
    """
    使用 API 调用模型，期望模型输出 JSON 文本，并解析为 Python 对象。
    支持重试机制和增强的JSON解析容错。

    Args:
        prompt: 提示词
        model: 模型名称
        max_output_tokens: 最大输出token数
        max_retries: 最大重试次数（默认3次）

    Returns:
        Any: 解析后的JSON对象
    """
    client = OpenAI(api_key=OPENAI_API_KEY, base_url=OPENAI_BASE_URL, timeout=180.0)

    last_error = None

    for attempt in range(max_retries):
        try:
            # 增加LLM调用计数（只在第一次调用时计数，重试不重复计数）
            if attempt == 0:
                call_count = increment_llm_call_count()
                print(f"  \033[1;36;40m[LLM调用 #{call_count}]\033[0m 正在调用模型: {model}")
            else:
                # 重试时显示重试信息
                logger.info(f"  [重试 {attempt}/{max_retries-1}] 重新调用API...")
                time.sleep(min(2 ** attempt, 8))  # 指数退避，最多8秒

            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7,
                max_tokens=max_output_tokens,
                stream=False
            )

            # 检查响应是否有效
            if not response or not response.choices or len(response.choices) == 0:
                raise ValueError("API返回空响应")

            content = response.choices[0].message.content
            if not content:
                raise ValueError("API返回内容为空")

            content = content.strip()
            if not content:
                raise ValueError("API返回内容为空字符串")

            # 预处理响应内容
            # 移除BOM标记
            if content.startswith('\ufeff'):
                content = content[1:]
            # 统一换行符
            content = content.replace('\r\n', '\n').replace('\r', '\n')
            # 移除不可见字符（保留换行和制表符）
            content = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', content)

            def _extract_balanced_json(txt: str) -> Optional[str]:
                """从文本中提取首个平衡的 JSON 块（{...} 或 [...]）"""
                start_idx = None
                start_ch = None
                for i, ch in enumerate(txt):
                    if ch in ["{", "["]:
                        start_idx = i
                        start_ch = ch
                        break
                if start_idx is None:
                    return None

                stack = 0
                closing = "}" if start_ch == "{" else "]"
                for j in range(start_idx, len(txt)):
                    c = txt[j]
                    if c == start_ch:
                        stack += 1
                    elif c == closing:
                        stack -= 1
                        if stack == 0:
                            return txt[start_idx:j + 1]
                return None

            # 策略1：尝试提取平衡的JSON
            json_str = _extract_balanced_json(content) or content

            # 策略2：清理控制字符后解析
            try:
                json_str_clean = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', json_str)
                return json.loads(json_str_clean)
            except Exception as e1:
                logger.debug(f"[JSON] 策略1失败: {e1}, 内容前100字符: {json_str[:100]}")

                # 策略3：移除所有控制字符（保留换行和制表符）
                try:
                    json_str_clean2 = ''.join(
                        ch for ch in json_str if ord(ch) >= 32 or ch in '\n\r\t'
                    )
                    return json.loads(json_str_clean2)
                except Exception as e2:
                    logger.debug(f"[JSON] 策略2失败: {e2}, 内容前100字符: {json_str[:100]}")

                    # 策略4：尝试修复常见JSON错误
                    try:
                        # 移除尾逗号
                        json_str_fixed = re.sub(r',\s*}', '}', json_str_clean2)
                        json_str_fixed = re.sub(r',\s*]', ']', json_str_fixed)
                        # 尝试添加缺失的引号（简单情况）
                        return json.loads(json_str_fixed)
                    except Exception as e3:
                        logger.debug(f"[JSON] 策略3失败: {e3}")

                        # 策略5：尝试从文本中提取关键信息
                        extracted = _extract_json_from_text(content)
                        if extracted:
                            logger.warning(f"[JSON] 使用提取的关键信息，原始内容长度: {len(content)}")
                            return extracted

                        # 所有策略都失败
                        if attempt < max_retries - 1 and _is_retryable_error(e3):
                            last_error = e3
                            continue  # 重试
                        else:
                            raise ValueError(f"JSON解析失败（所有策略都失败）: {e3}, 原始内容长度: {len(content)}, 前200字符: {content[:200]}")

        except Exception as e:
            last_error = e
            error_msg = str(e)

            # 判断是否可重试
            if attempt < max_retries - 1 and _is_retryable_error(e):
                logger.warning(f"API调用失败（可重试）: {error_msg[:100]}, 将在 {min(2 ** attempt, 8)} 秒后重试...")
                continue
            else:
                # 不可重试或已达到最大重试次数
                if "Expecting value" in error_msg or "JSON" in error_msg:
                    logger.warning(f"API调用失败（JSON解析错误）: {error_msg[:100]}")
                else:
                    logger.error(f"API调用失败: {error_msg[:100]}")
                raise

    # 所有重试都失败
    if last_error:
        raise last_error
    else:
        raise Exception("API调用失败：未知错误")

# --- 章节拆分处理器 ---
class BookProcessor:
    def __init__(self):
        self.chinese_patterns = [
            re.compile(r'^第[一二三四五六七八九十百千万]+章[：:：\s]*(.+)?'),
            re.compile(r'^第[0-9]+章[：:：\s]*(.+)?'),
            re.compile(r'^第[一二三四五六七八九十百千万]+节[：:：\s]*(.+)?'),
            re.compile(r'^第[0-9]+节[：:：\s]*(.+)?'),
            re.compile(r'^[一二三四五六七八九十百千万]+[、．.]\s*(.+)?'),
            re.compile(r'^[0-9]+[、．.]\s*(.+)?'),
        ]
        self.english_patterns = [
            re.compile(r'^Chapter\s+([0-9]+(?:\.[0-9]+)*)[：:：\s]*(.+)?', re.IGNORECASE),
            re.compile(r'^Chapter\s+([IVX]+)[：:：\s]*(.+)?', re.IGNORECASE),
            re.compile(r'^Section\s+([0-9]+(?:\.[0-9]+)*)[：:：\s]*(.+)?', re.IGNORECASE),
            re.compile(r'^Part\s+[0-9IVX]+[：:：\s]*(.+)?', re.IGNORECASE),
            re.compile(r'^([0-9]+(?:\.[0-9]+)*)[\.\s]+(.+)?'),
        ]
        self.markdown_header_pattern = re.compile(r'^(#{1,6})\s+(.+)$')


    def _detect_numeric_level(self, raw_title: str) -> Optional[int]:
        """
        检测数字层级格式的标题级别（如 1, 1.1, 1.1.1）
        返回层级深度：1表示第一级，2表示第二级，以此类推
        如果无法识别，返回None
        """
        # 匹配数字层级格式：1, 1.1, 1.1.1, 1.1.2 等
        numeric_pattern = re.compile(r'^([0-9]+(?:\.[0-9]+)*)')
        match = numeric_pattern.match(raw_title.strip())
        if match:
            numeric_part = match.group(1)
            # 计算点号数量+1即为层级深度
            level = numeric_part.count('.') + 1
            return level
        return None

    def _clean_title_text(self, title: str, remove_numeric_prefix: bool = False) -> str:
        """
        清理标题文本：去除页码、省略号、括号等，保留内容
        Args:
            title: 原始标题文本
            remove_numeric_prefix: 是否去除开头的数字前缀
        Returns:
            清理后的标题文本
        """
        title_clean = title.strip()
        # 去除 >> 前缀（如果存在）
        title_clean = re.sub(r'^>>\s*', '', title_clean).strip()
        if remove_numeric_prefix:
            # 先尝试匹配有空格的情况：1.1 标题
            title_clean = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s+', '', title_clean).strip()
            # 如果还有数字前缀（无空格的情况：1绪论），也去除
            title_clean = re.sub(r'^([0-9]+(?:\.[0-9]+)*)([^\d\s])', r'\2', title_clean).strip()
        title_clean = re.sub(r'\s+[0-9]+\s*$', '', title_clean).strip()  # 去除行末页码
        title_clean = re.sub(r'[．…]+.*$', '', title_clean).strip()  # 去除省略号后的内容
        title_clean = re.sub(r'[（(]([^）)]*)[）)]', r'\1', title_clean).strip()  # 只删除括号，保留内容
        title_clean = re.sub(r'\s+', ' ', title_clean).strip()  # 规范化空格
        return title_clean

    def _check_end_section_keyword(self, line_stripped: str) -> Optional[str]:
        """检查是否遇到结束标记关键词"""
        is_title_format = (
            self.markdown_header_pattern.match(line_stripped) is not None or
            any(pattern.match(line_stripped) for pattern in self.chinese_patterns + self.english_patterns)
        )

        if not is_title_format:
            return None

        for end_keyword in END_SECTION_KEYWORDS:
            title_text = line_stripped
            header_match = self.markdown_header_pattern.match(line_stripped)
            if header_match:
                title_text = header_match.group(2).strip()
            else:
                for pattern in self.chinese_patterns + self.english_patterns:
                    match = pattern.match(line_stripped)
                    if match:
                        if len(match.groups()) > 1 and match.group(2):
                            title_text = match.group(2).strip()
                        elif len(match.groups()) > 0 and match.group(1):
                            title_text = match.group(1).strip()
                        break
                title_text = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s+', '', title_text).strip()

            if end_keyword in title_text:
                return end_keyword
        return None

    def _extract_numeric_prefix(self, raw_title: str) -> Optional[str]:
        """
        提取标题中的数字前缀（如从"1.1 标题"中提取"1.1"）
        也支持中文章节标识：如"第一章"、"第一节"等
        支持实验编号格式：如"实验3-1"、"生物信息学 3-1"、"植物学 3-2"中提取"3-1"、"3-2"等
        """
        title_clean = raw_title.strip()
        title_clean = re.sub(r'^#+\s*', '', title_clean)
        title_clean = re.sub(r'^>>\s*', '', title_clean)  # 去除 >> 前缀

        # 先尝试提取数字前缀（如1.1, 1.1.2）
        title_no_space = re.sub(r'\s+', '', title_clean)
        numeric_pattern = re.compile(r'^([0-9]+(?:\.[0-9]+)*)')
        match = numeric_pattern.match(title_no_space)
        if match:
            return match.group(1)

        # 尝试提取通用编号格式（如"实验3-1"、"生物信息学 3-1"、"植物学 3-2"中的"3-1"、"3-2"）
        # 匹配模式：任意中文字符+可选空格+数字-数字
        # 例如：实验3-1、生物信息学 3-1、植物学 3-2
        general_number_pattern = re.compile(r'[\u4e00-\u9fa5]+\s*([0-9]+-[0-9]+)')
        gen_match = general_number_pattern.search(title_clean)
        if gen_match:
            return gen_match.group(1)

        # 尝试提取中文章节标识（如"第一章"、"第一节"）
        chinese_chapter_match = re.match(r'^(第[一二三四五六七八九十百千万0-9]+[章节])', title_clean)
        if chinese_chapter_match:
            return chinese_chapter_match.group(1)

        return None

    def _extract_core_title(self, title: str) -> str:
        """
        从标题中提取核心文本部分（去除前缀、页码、省略号等）

        例如：
        - "第一章 贾昕晔" -> "贾昕晔"
        - "1.1 贾昕晔" -> "贾昕晔"
        - "第一章 贾昕晔 (123)" -> "贾昕晔"
        - "1.1 贾昕晔 ………………………………………… (123)" -> "贾昕晔"
        - "实验3-1 常用实验样品的收集制备" -> "常用实验样品的收集制备"
        - "# >>实验3-1 常用实验样品的收集制备" -> "常用实验样品的收集制备"

        Args:
            title: 原始标题

        Returns:
            核心标题文本（去除所有前缀、页码、省略号、空格等）
        """
        # 去除markdown标记和 >> 前缀
        title_clean = re.sub(r'^#+\s*', '', title.strip())
        title_clean = re.sub(r'^>>\s*', '', title_clean)

        # 去除页码+章节编号格式（如"0011 1 "、"003 1.1.1 "等）
        # 匹配模式：数字（页码）+ 空格 + 数字或数字.数字（章节编号）+ 空格
        title_clean = re.sub(r'^[0-9]+\s+[0-9]+(?:\.[0-9]+)*\s+', '', title_clean)

        # 去除通用编号前缀（如"实验3-1 "、"生物信息学 3-1 "、"植物学 3-2 "等）
        # 匹配模式：任意中文字符+可选空格+数字-数字+空格
        title_clean = re.sub(r'^[\u4e00-\u9fa5]+\s*[0-9]+-[0-9]+\s+', '', title_clean)

        # 去除数字前缀（如"1.1 "、"1.1.1 "等）
        title_clean = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s+', '', title_clean)

        # 去除中文章节前缀（如"第一章 "、"第一节 "、"第二节 "等）
        # 注意：要匹配"章"和"节"
        title_clean = re.sub(r'^第[一二三四五六七八九十百千万0-9]+[章节]\s+', '', title_clean)

        # 去除页码（如"(123)"、"123"）
        title_clean = re.sub(r'\s*\([0-9]+\)\s*$', '', title_clean)
        title_clean = re.sub(r'\s+[0-9]+\s*$', '', title_clean)

        # 去除空格后的人名、页码等内容（如" 杨洪全/593"、" 作者名"、" /123"等）
        # 匹配模式：空格 + 可选的中文姓名 + 可选的"/" + 可选的数字
        # 例如：" 杨洪全/593"、" 张三"、" /123"、" 李四/456"
        title_clean = re.sub(r'\s+[\u4e00-\u9fa5]+(?:\/[0-9]+)?\s*$', '', title_clean)
        title_clean = re.sub(r'\s+\/[0-9]+\s*$', '', title_clean)  # 处理" /123"这种格式
        title_clean = re.sub(r'\s+[0-9]+\s*$', '', title_clean)  # 处理" 123"这种格式

        # 去除省略号
        title_clean = re.sub(r'…+', '', title_clean)
        title_clean = re.sub(r'\.{3,}', '', title_clean)

        # 去除多余空格
        title_clean = re.sub(r'\s+', '', title_clean)

        # 去除所有标点符号和特殊字符，只保留中文字符、英文字母和数字
        # 包括：·、•、。、，、：、；、！、？、.、,、:、;、!、?等
        title_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', title_clean)

        return title_clean

    def _is_title_only_line(self, line_stripped: str, title_info: Dict[str, Any]) -> bool:
        """
        检查该行是否只包含标题内容，无其他文字

        标题行的特征：
        1. 只包含标题文本（可能包含前缀如"# "、"3.2.2 "等）
        2. 可能包含页码（行末的数字）
        3. 可能包含省略号等格式字符
        4. 不应该包含其他描述性文字

        Args:
            line_stripped: 去除首尾空白的行
            title_info: 从该行提取的标题信息

        Returns:
            bool: 如果该行只包含标题内容，返回True；否则返回False
        """
        if not title_info or not title_info.get('is_title'):
            return False

        # 提取核心文本（去除前缀、页码等）
        core_title = self._extract_core_title(line_stripped)
        if not core_title:
            return False

        # 去除markdown标记
        line_no_markdown = re.sub(r'^#+\s*', '', line_stripped).strip()
        if not line_no_markdown:
            return False

        # 去除数字前缀（如"3.2.2 "、"1.1.1 "等）
        text_after_prefix = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s+', '', line_no_markdown)

        # 去除中文章节前缀（如"第一章 "、"第一节 "等）
        text_after_chapter = re.sub(r'^第[一二三四五六七八九十百千万0-9]+[章节]\s+', '', text_after_prefix)

        # 去除通用编号前缀（如"实验3-1 "等）
        text_after_general = re.sub(r'^[\u4e00-\u9fa5]+\s*[0-9]+-[0-9]+\s+', '', text_after_chapter)

        # 去除页码（行末的数字，如" 56"）
        text_after_page = re.sub(r'\s+[0-9]+\s*$', '', text_after_general)

        # 去除省略号等格式字符
        text_after_format = re.sub(r'[．…\.]{2,}', '', text_after_page)
        text_after_format = re.sub(r'[·•]', '', text_after_format)

        # 去除所有空格和标点符号，只保留字符
        text_clean = re.sub(r'\s+', '', text_after_format)
        text_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', text_clean)

        # core_title 已经通过 _extract_core_title 去除了所有标点符号
        core_title_clean = core_title  # _extract_core_title 已经去除了所有标点符号

        # 如果清理后的文本完全等于核心标题文本，肯定是标题行
        if text_clean == core_title_clean:
            return True

        # 如果核心标题文本在清理后的文本中，且占比足够高
        if core_title_clean and text_clean and core_title_clean in text_clean:
            title_ratio = len(core_title_clean) / len(text_clean) if text_clean else 0
            if title_ratio >= 0.7:
                return True

        # 对于markdown标题，如果整行匹配标题模式，进一步检查
        if title_info.get('is_markdown'):
            markdown_pattern = r'^#+\s+.+$'
            if re.match(markdown_pattern, line_stripped):
                # 去除markdown标记和可能的页码后，检查是否只包含标题文本
                title_without_markdown = re.sub(r'^#+\s+', '', line_stripped).strip()
                title_without_page = re.sub(r'\s+[0-9]+\s*$', '', title_without_markdown)
                title_without_page_clean = re.sub(r'\s+', '', title_without_page)
                title_without_page_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', title_without_page_clean)

                if core_title_clean and title_without_page_clean:
                    if core_title_clean in title_without_page_clean:
                        ratio = len(core_title_clean) / len(title_without_page_clean) if title_without_page_clean else 0
                        if ratio >= 0.7:
                            return True

        return False

    def _is_numeric_parent_of(self, parent_num: str, child_num: str) -> bool:
        """
        判断parent_num是否是child_num的父级
        例如：1 是 1.1 的父级，1.1 是 1.1.1 的父级
        """
        if not parent_num or not child_num:
            return False
        # 父级编号应该是子级编号的前缀（以点号分隔）
        return child_num.startswith(parent_num + '.')

    def _extract_title_from_line(self, line_stripped: str) -> Optional[Dict[str, Any]]:
        """
        从一行文本中提取标题信息（公共方法，避免代码重复）

        Returns:
            {
                'is_title': bool,
                'title_clean': str,
                'line_numeric': Optional[str],
                'level': Optional[int],
                'is_markdown': bool
            } 或 None（如果不是标题）
        """
        if not line_stripped:
            return None

        # 检查markdown标题
        header_match = self.markdown_header_pattern.match(line_stripped)
        if header_match:
            title = header_match.group(2).strip()
            # 对于markdown标题，先清理标题文本（去除页码等），但保留前缀
            # 前缀会在 _extract_core_title 中去除
            title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
            line_numeric = self._extract_numeric_prefix(line_stripped)
            level = len(header_match.group(1))
            return {
                'is_title': True,
                'title_clean': title_clean,
                'line_numeric': line_numeric,
                'level': level,
                'is_markdown': True
            }

        # 检查纯数字前缀格式（如 "1.1 标题"）
        numeric_prefix_match = re.match(r'^([0-9]+(?:\.[0-9]+)*)\s+(.+)$', line_stripped)
        if numeric_prefix_match:
            numeric_prefix = numeric_prefix_match.group(1)
            title = numeric_prefix_match.group(2).strip()
            title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
            numeric_level = self._detect_numeric_level(line_stripped)
            level = numeric_level if numeric_level is not None else 2
            return {
                'is_title': True,
                'title_clean': title_clean,
                'line_numeric': numeric_prefix,
                'level': level,
                'is_markdown': False
            }

        # 检查通用编号格式（如"实验3-1"、"生物信息学 3-1"、"植物学 3-2"等）
        # 匹配模式：任意中文字符+可选空格+数字-数字+空格+标题
        general_number_match = re.match(r'^([\u4e00-\u9fa5]+)\s*([0-9]+-[0-9]+)\s+(.+)$', line_stripped)
        if general_number_match:
            subject_name = general_number_match.group(1)  # 学科名
            numeric_prefix = general_number_match.group(2)  # 编号
            title = general_number_match.group(3).strip()  # 标题文本
            title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
            # 通用编号格式通常作为3级标题
            level = 3
            # 保留完整标题格式：学科名编号 标题
            full_title = f"{subject_name}{numeric_prefix} {title_clean}".strip()
            return {
                'is_title': True,
                'title_clean': full_title,
                'line_numeric': numeric_prefix,
                'level': level,
                'is_markdown': False
            }

        # 检查章节标题模式
        for pattern in self.chinese_patterns + self.english_patterns:
            match = pattern.match(line_stripped)
            if match:
                # 优先使用第二个group（标题文本），如果没有则使用第一个group
                if len(match.groups()) > 1 and match.group(2):
                    title = match.group(2).strip()
                elif len(match.groups()) > 0 and match.group(1):
                    title = match.group(1).strip()
                else:
                    title = line_stripped

                if not title:
                    title = line_stripped

                title_clean = self._clean_title_text(title)
                line_numeric = self._extract_numeric_prefix(line_stripped)

                # 判断级别
                numeric_level = self._detect_numeric_level(line_stripped)
                if numeric_level is not None:
                    level = numeric_level
                elif '节' in line_stripped:
                    level = 2
                elif '章' in line_stripped:
                    level = 1
                else:
                    level = 2

                return {
                    'is_title': True,
                    'title_clean': title_clean,
                    'line_numeric': line_numeric,
                    'level': level,
                    'is_markdown': False
                }

        return None

    def _is_in_sub_toc(self, lines: List[str], current_line_idx: int, look_back: int = 40, look_ahead: int = 40) -> bool:
        """
        检查当前位置是否在小目录区域内

        小目录的特征（基于实际文件分析）：
        1. 标题密度非常高（>70%）
        2. 连续标题很多（>=10个）
        3. 标题之间只有空行或很短的文本
        4. 小目录后面会有真正的正文标题和正文内容

        Args:
            lines: 所有行
            current_line_idx: 当前行索引
            look_back: 向前查看的行数
            look_ahead: 向后查看的行数

        Returns:
            bool: 是否在小目录区域内
        """
        # 首先检查当前行是否在目录区域内（通过检测目录区域）
        # 如果当前行在目录区域之后，且是markdown格式的章节标题，不应该被认为是小目录
        # 因为这是正文开始处的标题
        try:
            toc_start, toc_end = self._detect_toc_without_title(lines)
            if toc_start != -1 and toc_end != -1:
                # 如果当前行在目录区域之后，且是markdown格式的章节标题
                if current_line_idx > toc_end:
                    line_stripped = lines[current_line_idx].strip()
                    if line_stripped.startswith('#'):
                        title_info = self._extract_title_from_line(line_stripped)
                        if title_info and title_info.get('is_title'):
                            # 检查是否是章节标题（包含"第X章"、"第X节"等，或者是单独的标题）
                            title_clean = title_info.get('title_clean', '')
                            # 如果是"第X章"格式，或者是单独的标题（没有页码），很可能是正文开始
                            if re.search(r'^第[一二三四五六七八九十百千万0-9]+[章节]', title_clean) or \
                               (not re.search(r'\s+[0-9]+\s*$', line_stripped) and len(title_clean) > 2):
                                # 这是正文开始处的标题，不应该被认为是小目录
                                return False
        except:
            pass  # 如果检测失败，继续使用原有逻辑
        # 方法1: 检查是否包含明显的目录关键词
        sub_toc_keywords = ['## 目录', '## Contents', '## 本章目录', '## 本章内容',
                           '## 章节目录', '## 内容提要', '## 本章要点', '### 目录', '### Contents']

        # 向前查找目录关键词（最多向前30行）
        for i in range(max(0, current_line_idx - 30), current_line_idx + 1):
            line = lines[i].strip()
            for keyword in sub_toc_keywords:
                if keyword in line:
                    # 找到目录关键词，检查后续是否有连续的标题
                    title_count = 0
                    non_empty_count = 0
                    for j in range(i + 1, min(i + 40, len(lines))):
                        check_line = lines[j].strip()
                        if not check_line:
                            continue
                        non_empty_count += 1
                        title_info = self._extract_title_from_line(check_line)
                        if title_info and title_info['is_title']:
                            title_count += 1

                    # 如果标题密度超过30%，且当前行在关键词之后，认为是小目录
                    if non_empty_count > 0 and title_count / non_empty_count > 0.3:
                        if current_line_idx >= i:
                            # 进一步确认：检查后续是否有正文（超过60字符的非标题行）
                            has_body_after = False
                            for j in range(i + 1, min(i + 80, len(lines))):
                                check_line = lines[j].strip()
                                if not check_line:
                                    continue
                                if len(check_line) > 60:
                                    title_info = self._extract_title_from_line(check_line)
                                    if not title_info or not title_info['is_title']:
                                        has_body_after = True
                                        break

                            if has_body_after:
                                return True

        # 方法2: 检查标题密度（基于实际文件：标题密度>70%，连续标题>=10个）
        # 检查当前位置周围40行的标题密度
        start_idx = max(0, current_line_idx - 20)
        end_idx = min(len(lines), current_line_idx + 20)

        title_count = 0
        non_empty_count = 0
        consecutive_titles = 0
        max_consecutive = 0

        for i in range(start_idx, end_idx):
            line = lines[i].strip()
            if not line:
                continue
            non_empty_count += 1

            title_info = self._extract_title_from_line(line)
            if title_info and title_info['is_title']:
                # 只统计真正的标题行（只包含标题内容，无其他文字）
                is_title_only = self._is_title_only_line(line, title_info)
                if is_title_only:
                    title_count += 1
                    consecutive_titles += 1
                    max_consecutive = max(max_consecutive, consecutive_titles)
            else:
                consecutive_titles = 0

        # 如果标题密度很高（>60%）且连续标题很多（>=8个），可能是小目录
        if non_empty_count > 0:
            title_density = title_count / non_empty_count
            if title_density > 0.6 and max_consecutive >= 8:
                # 进一步确认：检查后续是否有正文（超过60字符的非标题行）
                has_body_after = False
                for i in range(current_line_idx + 1, min(current_line_idx + 80, len(lines))):
                    line = lines[i].strip()
                    if not line:
                        continue
                    if len(line) > 60:
                        title_info = self._extract_title_from_line(line)
                        if not title_info or not title_info['is_title']:
                            has_body_after = True
                            break

                # 如果后面有正文，且当前位置标题密度高，可能是小目录
                # 但是，如果当前标题是正文开始处的标题（前面没有很多标题），不应该认为是小目录
                # 检查前面是否有足够的标题（至少5个）来确认这是小目录区域
                titles_before = 0
                for i in range(max(0, current_line_idx - 30), current_line_idx):
                    line = lines[i].strip()
                    if not line:
                        continue
                    title_info = self._extract_title_from_line(line)
                    if title_info and title_info['is_title']:
                        is_title_only = self._is_title_only_line(line, title_info)
                        if is_title_only:
                            titles_before += 1

                # 如果前面有足够的标题（>=5个），且满足其他条件，才认为是小目录
                if has_body_after and titles_before >= 5:
                    return True

        return False

    def _detect_toc_without_title(self, lines: List[str], start_search: int = 0, max_search: int = 500) -> tuple[int, int]:
        """
        检测没有"目录"标题的目录区域

        目录特征：
        1. 高密度的标题行（连续多行都是标题格式）
        2. 标题行后面通常有页码（数字）
        3. 目录区域结束后，会有明显的正文开始标志

        Args:
            lines: 所有行
            start_search: 开始搜索的行号
            max_search: 最大搜索行数（避免搜索整个文件）

        Returns:
            (toc_start, toc_end): 目录开始和结束行号，如果未找到返回(-1, -1)
        """
        # 目录结束标志
        toc_end_keywords = ['中英文名词对照索引', '参考文献', '主要参考文献', 'References', 'REFERENCES',
                           '附录', 'Appendix', 'APPENDIX', '索引', 'Index']

        # 从start_search开始，最多搜索max_search行
        search_end = min(start_search + max_search, len(lines))

        # 查找第一个标题行作为目录开始候选
        toc_start_candidate = -1
        for i in range(start_search, search_end):
            line_stripped = lines[i].strip()
            if not line_stripped:
                continue

            # 检查是否是标题（markdown标题或章节标题）
            title_info = self._extract_title_from_line(line_stripped)
            if title_info and title_info['is_title']:
                # 跳过明显的非目录标题（如"前言"、"序"等，但保留"第一篇"、"第X章"等）
                title_clean = title_info['title_clean']
                # 如果是"第X篇"、"第X章"等格式，很可能是目录开始
                if re.search(r'^第[一二三四五六七八九十百千万0-9]+[篇章]', line_stripped) or \
                   re.search(r'^#\s*第[一二三四五六七八九十百千万0-9]+[篇章]', line_stripped):
                    toc_start_candidate = i
                    break
                # 或者如果标题密度足够高，也可能是目录
                # 先记录，继续检查后续行的标题密度
                if toc_start_candidate == -1:
                    toc_start_candidate = i

        if toc_start_candidate == -1:
            return (-1, -1)

        # 从候选开始位置，检查后续区域的标题密度
        window_size = 100  # 检查窗口大小
        min_title_density = 0.25  # 最小标题密度（25%）
        min_toc_length = 15  # 最小目录长度（行数）

        # 统计从候选位置开始的标题密度
        title_count = 0
        total_non_empty = 0
        page_number_count = 0

        window_end = min(toc_start_candidate + window_size, search_end)
        for i in range(toc_start_candidate, window_end):
            line_stripped = lines[i].strip()
            if not line_stripped:
                continue

            total_non_empty += 1

            # 检查是否是标题
            title_info = self._extract_title_from_line(line_stripped)
            if title_info and title_info['is_title']:
                title_count += 1
                # 检查是否有页码（行末的数字）
                if re.search(r'\s+[0-9]+\s*$', line_stripped):
                    page_number_count += 1

        # 如果标题密度不够，不是目录
        if total_non_empty == 0 or title_count < min_toc_length:
            return (-1, -1)

        title_density = title_count / total_non_empty
        if title_density < min_title_density:
            return (-1, -1)

        # 找到目录开始位置（第一个标题行）
        toc_start = toc_start_candidate

        # 查找目录结束位置
        toc_end = -1
        # 从窗口结束位置继续向后查找
        # 优先查找markdown格式的结束标志（如"# 参考文献"）
        for i in range(window_end, min(window_end + 300, len(lines))):
            line_stripped = lines[i].strip()
            if not line_stripped:
                continue

            # 优先检查markdown格式的结束标志（如"# 参考文献"）
            if line_stripped.startswith('#'):
                for end_keyword in toc_end_keywords:
                    if end_keyword in line_stripped:
                        # 确认这是markdown标题格式的结束标志
                        toc_end = i
                        break
                if toc_end != -1:
                    break

            # 检查是否遇到目录结束标志（非markdown格式）
            if toc_end == -1:
                for end_keyword in toc_end_keywords:
                    if end_keyword in line_stripped:
                        # 但需要确保不在目录开始之前（避免匹配到前言中的"参考文献"）
                        # 对于"主要参考文献"等，即使不是markdown格式，也应该识别
                        if i > toc_start + 50:  # 确保在目录区域内
                            toc_end = i
                            break

            if toc_end != -1:
                break

            # 检查是否遇到正文开始标志
            # 1. 检查是否是markdown格式的章节标题（如"# 第一章"、"# 植物生产与环境"）
            # 如果遇到这种标题，且前面有足够的目录内容，说明目录已经结束
            if line_stripped.startswith('#'):
                title_info = self._extract_title_from_line(line_stripped)
                if title_info and title_info.get('is_title'):
                    # 检查是否是章节标题（包含"第X章"、"第X节"等，或者是单独的标题）
                    title_clean = title_info.get('title_clean', '')
                    # 如果是"第X章"格式，或者是单独的标题（没有页码），很可能是正文开始
                    if re.search(r'^第[一二三四五六七八九十百千万0-9]+[章节]', title_clean) or \
                       (not re.search(r'\s+[0-9]+\s*$', line_stripped) and i > toc_start + min_toc_length):
                        # 检查前面是否有足够的目录内容
                        # 如果当前标题前面有足够的目录行，且当前标题不在目录区域内，说明目录已结束
                        # 目录结束位置应该在当前标题之前
                        if i > toc_start + min_toc_length:
                            # 向前查找最后一个目录行（通常是有页码的标题行）
                            for j in range(i - 1, max(toc_start, i - 50), -1):
                                prev_line = lines[j].strip()
                                if not prev_line:
                                    continue
                                # 如果找到有页码的标题行，或者找到明显的目录结束标志
                                if re.search(r'\s+[0-9]+\s*$', prev_line):
                                    toc_end = j + 1  # 目录结束在最后一个有页码的行之后
                                    break
                            if toc_end == -1:
                                toc_end = i  # 如果没找到，目录结束在当前标题之前
                            break

            # 2. 检查是否遇到大段正文（非标题的长行，且不是图片）
            if len(line_stripped) > 100 and not line_stripped.startswith('!['):
                title_info = self._extract_title_from_line(line_stripped)
                if not title_info or not title_info['is_title']:
                    # 找到正文开始，目录结束在当前行之前
                    # 但需要确认前面确实有足够的标题
                    if i > toc_start + min_toc_length:
                        toc_end = i
                        break

        if toc_end == -1:
            # 如果没找到明确的结束标志，使用窗口结束位置
            toc_end = window_end

        return (toc_start, toc_end)

    def _split_from_body(self, lines: List[str], search_start: int) -> List[Dict[str, str]]:
        """
        从正文中直接提取章节（当目录中未找到标题时）

        Args:
            lines: 文本行列表
            search_start: 开始搜索的行号

        Returns:
            章节列表
        """
        logger.info("从正文中搜索章节标题")
        headers = []

        # 在正文中搜索所有标题
        for i in range(search_start, len(lines)):
            line_stripped = lines[i].strip()
            if not line_stripped:
                continue

            title_info = self._extract_title_from_line(line_stripped)
            if title_info and title_info['is_title']:
                # 跳过"目录"等非章节标题
                title_clean = title_info['title_clean']
                if title_clean.strip() in ["目录", "Contents", "CONTENTS", "目  录", "目　录"]:
                    continue

                headers.append({
                    'line_num': i,
                    'level': title_info['level'],
                    'title': title_clean,
                    'raw_title': line_stripped,
                    'numeric_prefix': title_info['line_numeric'],
                    'is_markdown': title_info['is_markdown']
                })

        if not headers:
            logger.warning("正文中未找到任何章节标题，使用回退策略")
            return self._fallback_split('\n'.join(lines[search_start:]))

        # 按标题分割章节
        final_chunks = []
        global_counter = 1

        for idx, header in enumerate(headers):
            start_line = header['line_num']
            end_line = len(lines)

            # 查找下一个标题
            if idx + 1 < len(headers):
                end_line = headers[idx + 1]['line_num']

            # 提取章节内容
            if end_line <= start_line:
                continue

            if start_line + 1 >= len(lines):
                continue

            section_lines = lines[start_line + 1:end_line]
            section_content = '\n'.join(section_lines)
            section_content = clean_text_basic(section_content)

            # 不再过滤短章节，提取所有章节

            chunk_id = f"chunk_{idx+1:03d}_{global_counter:03d}"
            chunk_title = header['title']

            final_chunks.append({
                'chunk_id': chunk_id,
                'chunk_title': chunk_title,
                'text': section_content,
                'level': header['level'],
                'parent_title': None
            })
            global_counter += 1

        logger.info(f"从正文中提取到 {len(final_chunks)} 个章节")
        return final_chunks

    def split_by_chapters(self, text: str) -> List[Dict[str, str]]:
        """拆分章节（从参考脚本中提取的核心方法）"""
        lines = text.split("\n")

        # 1. 提取目录信息（更精确地匹配）
        toc_keywords = ['目录', 'Contents', 'CONTENTS', '目  录', '目　录', '目 录']
        toc_start = -1
        toc_end = -1

        def is_toc_title(text: str) -> bool:
            """检查文本是否为目录标题（支持各种空格变体）"""
            if not text:
                return False
            # 去除所有空格后比较
            text_no_space = re.sub(r'\s+', '', text)
            toc_keywords_no_space = [re.sub(r'\s+', '', kw) for kw in toc_keywords]
            return text_no_space in toc_keywords_no_space or text_no_space == '目录'

        for i, line in enumerate(lines):
            line_stripped = line.strip()
            if toc_start == -1:
                # 优先匹配精确的"# 目录"格式
                header_match = self.markdown_header_pattern.match(line_stripped)
                if header_match:
                    title = header_match.group(2).strip()
                    # 使用灵活匹配：支持各种空格变体
                    if is_toc_title(title):
                        toc_start = i
                        # 不要break，继续检查目录结束
                # 如果没有markdown格式，检查是否是单独的"目录"行
                elif is_toc_title(line_stripped):
                    toc_start = i
                    # 不要break，继续检查目录结束
            elif toc_end == -1:
                # 目录结束判断：查找目录中的最后一个标题行
                # 当遇到标题行后，检查后续是否有大段正文，如果有则目录结束
                title_info = self._extract_title_from_line(line_stripped)
                if title_info and title_info['is_title']:
                    title_clean = title_info['title_clean']
                    # 跳过目录标题本身
                    if title_clean.strip() in toc_keywords:
                        continue

                    # 检查后续是否有大段正文（非标题行）
                    look_ahead = 10
                    has_body = False
                    for j in range(i + 1, min(i + 1 + look_ahead, len(lines))):
                        if j < len(lines):
                            next_line = lines[j].strip()
                            if next_line and len(next_line) > 100:
                                next_title_info = self._extract_title_from_line(next_line)
                                if not next_title_info or not next_title_info['is_title']:
                                    has_body = True
                                    break

                    if has_body:
                        # 找到正文开始，目录结束在当前标题行之后
                        toc_end = i + 1
                        break

        # 如果没有找到带"目录"标题的目录，尝试检测无标题的目录区域
        if toc_start == -1 or toc_end == -1:
            logger.info("未检测到带'目录'标题的目录区域，尝试检测无标题的目录区域")
            # 从前500行开始搜索（通常目录在文件前部）
            auto_toc_start, auto_toc_end = self._detect_toc_without_title(lines, start_search=0, max_search=500)
            if auto_toc_start != -1 and auto_toc_end != -1:
                toc_start = auto_toc_start
                toc_end = auto_toc_end
                logger.info(f"自动检测到目录区域：第 {toc_start+1} 行到第 {toc_end+1} 行")

        # 设置目录行和搜索起始位置
        if toc_start != -1 and toc_end != -1:
            toc_lines = lines[toc_start:toc_end]
            search_start = toc_end
        else:
            logger.info("未检测到明确的目录区域，将在全文中搜索章节标题")
            toc_lines = []
            search_start = 0

        # 2. 从目录中提取章节标题信息（只提取真正的章节标题，必须有前缀或后缀）
        toc_headers = []
        if toc_lines:
            for i, line in enumerate(toc_lines):
                line_stripped = line.strip()
                if not line_stripped:
                    continue

                # 跳过目录标题本身
                if line_stripped in ['# 目录', '目录', 'Contents', 'CONTENTS', '目  录', '目　录']:
                    continue

                header_match = self.markdown_header_pattern.match(line_stripped)
                if header_match:
                    level = len(header_match.group(1))
                    title = header_match.group(2).strip()

                    # 提取数字前缀和检查是否有章节标识
                    numeric_prefix = self._extract_numeric_prefix(line_stripped)
                    # 检查是否是通用编号格式（如"实验3-1"、"生物信息学 3-1"、"植物学 3-2"等）
                    # 匹配模式：任意中文字符+可选空格+数字-数字
                    is_general_number_format = bool(re.search(r'[\u4e00-\u9fa5]+\s*[0-9]+-[0-9]+', line_stripped))
                    has_chapter_marker = bool(
                        re.search(r'第[一二三四五六七八九十百千万0-9]+[章节]', line_stripped) or
                        re.search(r'Chapter\s+[0-9IVX]+', line_stripped, re.IGNORECASE) or
                        re.search(r'Section\s+[0-9]+', line_stripped, re.IGNORECASE) or
                        numeric_prefix or  # 有数字前缀（如1.1, 1.1.2, 3-1）
                        is_general_number_format  # 通用编号格式（学科名+数字-数字）
                    )

                    # 特殊处理：允许提取"绪论"、"前言"等常见无前缀标题
                    is_special_title = bool(
                        '绪论' in title or '前言' in title or 'Preface' in title or 'Introduction' in title
                    )

                    # 只提取有前缀或特殊标题（忽略"实验指导"、"主要参考文献"等）
                    if not has_chapter_marker and not is_special_title:
                        continue

                    # 先提取前缀
                    numeric_prefix = self._extract_numeric_prefix(line_stripped)

                    # 清理标题（如果前缀是中文章节标识如"第一章"，不要去除，因为它已经是标题的一部分）
                    if numeric_prefix and re.match(r'^第[一二三四五六七八九十百千万0-9]+[章节]', numeric_prefix):
                        # 中文章节标识，保留在标题中，只清理页码等
                        title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
                        final_title = title_clean
                    else:
                        # 数字前缀，去除后重新组合
                        title_clean = self._clean_title_text(title, remove_numeric_prefix=True)
                        if numeric_prefix:
                            final_title = f"{numeric_prefix} {title_clean}".strip()
                        else:
                            final_title = title_clean

                    numeric_level = self._detect_numeric_level(line_stripped)
                    if numeric_level is not None:
                        actual_level = numeric_level
                    elif '节' in final_title:
                        actual_level = 2
                    elif '章' in final_title:
                        actual_level = 1
                    else:
                        actual_level = level

                    # 只添加有意义的标题（不是纯页码、不是太短的标题）
                    if len(final_title) > 1 and not re.match(r'^[0-9\s\(\)]+$', final_title):
                        toc_headers.append({
                            'level': actual_level,
                            'title': final_title,
                            'raw_title': line_stripped,
                            'numeric_prefix': numeric_prefix,
                            'is_markdown': True,
                            'toc_line': i + toc_start
                        })
                else:
                    # 跳过纯页码行、省略号行等
                    if re.match(r'^[0-9\s\(\)…\.]+$', line_stripped) or len(line_stripped) < 2:
                        continue

                    # 检查是否有章节标识（前缀或后缀）
                    numeric_prefix = self._extract_numeric_prefix(line_stripped)
                    # 检查是否是通用编号格式（如"实验3-1"、"生物信息学 3-1"、"植物学 3-2"等）
                    # 匹配模式：任意中文字符+可选空格+数字-数字
                    is_general_number_format = bool(re.search(r'[\u4e00-\u9fa5]+\s*[0-9]+-[0-9]+', line_stripped))
                    has_chapter_marker = bool(
                        re.search(r'第[一二三四五六七八九十百千万0-9]+[章节]', line_stripped) or
                        re.search(r'Chapter\s+[0-9IVX]+', line_stripped, re.IGNORECASE) or
                        re.search(r'Section\s+[0-9]+', line_stripped, re.IGNORECASE) or
                        numeric_prefix or  # 有数字前缀（如1.1, 1.1.2, 3-1）
                        is_general_number_format  # 通用编号格式（学科名+数字-数字）
                    )

                    # 特殊处理：允许提取"绪论"、"前言"等常见无前缀标题
                    is_special_title = bool(
                        '绪论' in line_stripped or '前言' in line_stripped or
                        'Preface' in line_stripped or 'Introduction' in line_stripped
                    )

                    # 只提取有前缀或特殊标题
                    if not has_chapter_marker and not is_special_title:
                        continue

                    # 处理通用编号格式（如"实验3-1 常用实验样品的收集制备 025"、"生物信息学 3-1 标题"等）
                    # 匹配模式：任意中文字符+可选空格+数字-数字+空格+标题+可选页码
                    general_number_match = re.match(r'([\u4e00-\u9fa5]+)\s*([0-9]+-[0-9]+)\s+(.+?)(?:\s+[0-9]+)?\s*$', line_stripped)
                    if general_number_match:
                        subject_name = general_number_match.group(1)  # 学科名（如"实验"、"生物信息学"、"植物学"）
                        numeric_prefix = general_number_match.group(2)  # 编号（如"3-1"）
                        title = general_number_match.group(3).strip()  # 标题文本
                        # 清理标题（去除页码、省略号等）
                        title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
                        # 通用编号格式通常作为3级标题
                        level = 3
                        # 保留学科名和编号，格式：学科名编号 标题（如"实验3-1 常用实验样品的收集制备"）
                        final_title = f"{subject_name}{numeric_prefix} {title_clean}".strip()
                        # 只添加有意义的标题
                        if len(final_title) > 1:
                            toc_headers.append({
                                "level": level,
                                "title": final_title,
                                "raw_title": line_stripped,
                                "numeric_prefix": numeric_prefix,
                                "is_markdown": False,
                                "toc_line": i + toc_start
                            })
                        continue

                    numeric_prefix_match = re.match(r'^([0-9]+(?:\.[0-9]+)*)\s+(.+)$', line_stripped)
                    if numeric_prefix_match:
                        numeric_prefix = numeric_prefix_match.group(1)
                        title = numeric_prefix_match.group(2).strip()
                        # 清理标题（去除页码、省略号等）
                        title_clean = self._clean_title_text(title, remove_numeric_prefix=False)
                        numeric_level = self._detect_numeric_level(line_stripped)
                        if numeric_level is not None:
                            level = numeric_level
                        else:
                            level = 2
                        final_title = f"{numeric_prefix} {title_clean}".strip()
                        # 只添加有意义的标题
                        if len(final_title) > 1:
                            toc_headers.append({
                                "level": level,
                                "title": final_title,
                                "raw_title": line_stripped,
                                "numeric_prefix": numeric_prefix,
                                "is_markdown": False,
                                "toc_line": i + toc_start
                            })
                    else:
                        for pattern in self.chinese_patterns + self.english_patterns:
                            match = pattern.match(line_stripped)
                            if match:
                                if len(match.groups()) > 1 and match.group(2):
                                    title = match.group(2).strip()
                                elif len(match.groups()) > 0 and match.group(1):
                                    title = match.group(1).strip()
                                else:
                                    title = line_stripped

                                if not title:
                                    title = line_stripped

                                title_clean = self._clean_title_text(title, remove_numeric_prefix=True)

                                if '节' in line_stripped:
                                    level = 2
                                elif '章' in line_stripped:
                                    level = 1
                                else:
                                    numeric_level = self._detect_numeric_level(line_stripped)
                                    if numeric_level is not None:
                                        level = numeric_level
                                    else:
                                        level = 2

                                numeric_prefix = self._extract_numeric_prefix(line_stripped)
                                if numeric_prefix:
                                    final_title = f"{numeric_prefix} {title_clean}".strip()
                                else:
                                    final_title = title_clean

                                # 只添加有意义的标题
                                if len(final_title) > 1:
                                    toc_headers.append({
                                        "level": level,
                                        "title": final_title,
                                        "raw_title": line_stripped,
                                        "numeric_prefix": numeric_prefix,
                                        "is_markdown": False,
                                        "toc_line": i + toc_start
                                    })
                                break

        # 3. 严格按照目录中的标题在正文中查找并分割
        # 如果没有目录，返回空列表（不进行回退，严格按照目录）
        if not toc_headers:
            logger.warning("未找到目录，无法提取章节（严格按照目录提取）")
            return []

        # 过滤掉目录标题本身（如"目录"、"Contents"等）
        filtered_toc_headers = []
        for toc_header in toc_headers:
            title_clean = toc_header['title'].strip()
            # 跳过目录标题本身
            if title_clean in ["目录", "Contents", "CONTENTS", "目  录", "目　录"]:
                continue
            filtered_toc_headers.append(toc_header)

        if not filtered_toc_headers:
            logger.warning("目录中只有目录标题本身，没有实际章节标题")
            return []

        # 在正文中按顺序查找目录中的每个标题，并记录其行号
        # 关键：必须按照目录顺序，从前到后依次查找，每次从上次找到的位置之后开始
        toc_title_positions = []  # [(line_num, toc_header), ...]
        last_found_line = search_start - 1  # 上次找到的位置，初始为目录结束位置之前

        for toc_header in filtered_toc_headers:
            title_to_find = toc_header['title']
            numeric_prefix = toc_header.get('numeric_prefix')

            # 提取目录标题的核心文本（去除前缀、页码、省略号等）
            core_title_to_find = self._extract_core_title(title_to_find)

            # 在正文中查找该标题（从上次找到的位置之后开始，确保按顺序）
            # 重要：必须从目录结束位置之后开始搜索，不提取目录前的内容
            found_line = -1
            search_from = max(search_start, last_found_line + 1)  # 从目录结束位置或上次找到的位置之后开始

            for i in range(search_from, len(lines)):
                # 确保不在目录区域内（双重保险）
                if toc_start != -1 and toc_end != -1 and toc_start <= i < toc_end:
                    continue  # 跳过目录区域内的内容

                line_stripped = lines[i].strip()
                if not line_stripped:
                    continue

                # 检查该行是否是标题（独立一行）
                title_info = self._extract_title_from_line(line_stripped)
                if not title_info or not title_info['is_title']:
                    continue

                # 检查该行是否只包含标题内容，无其他文字
                if not self._is_title_only_line(line_stripped, title_info):
                    continue  # 如果该行包含其他文字内容，跳过

                # 提取正文中标题的核心文本
                # 重要：直接使用原始行文本提取核心文本，确保能正确去除所有前缀
                # 因为 _extract_core_title 会处理 markdown 标记、前缀等所有格式
                line_core_title = self._extract_core_title(line_stripped)

                # 严格匹配：只匹配核心文本部分（忽略前缀、页码、省略号等）
                # 核心文本必须完全匹配（只比较字符，忽略标点符号和空格）
                match_success = False
                if core_title_to_find and line_core_title:
                    # 由于 _extract_core_title 已经去除了所有标点符号和前缀，直接比较即可
                    # 但为了保险，再次去除标点符号和空格进行比较（只保留中文字符、英文字母和数字）
                    core1_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', core_title_to_find)
                    core2_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', line_core_title)

                    # 完全匹配（忽略空格和标点符号）
                    if core1_clean == core2_clean and len(core1_clean) >= 1:
                        match_success = True
                    # 如果完全匹配失败，尝试模糊匹配（对于短标题，允许部分匹配）
                    elif len(core1_clean) >= 2 and len(core2_clean) >= 2:
                        # 如果核心文本长度>=2，且一个包含另一个，且长度差异不大，认为是匹配
                        if (core1_clean in core2_clean or core2_clean in core1_clean) and abs(len(core1_clean) - len(core2_clean)) <= 2:
                            match_success = True

                # 如果匹配成功，检查是否在小目录区域
                if match_success:
                    # 检查匹配到的标题是否在小目录区域
                    # 小目录区域标志：连续的标题，无其他文本内容
                    is_in_sub_toc = self._is_in_sub_toc(lines, i)
                    if is_in_sub_toc:
                        # 如果在小目录区域，跳过这个匹配，继续查找下一个匹配
                        continue

                    # 如果不在小目录区域，接受匹配
                    found_line = i
                    break

            # 如果严格匹配失败，尝试模糊匹配（对于"第X章 标题"格式，尝试只匹配标题部分）
            if found_line == -1 and core_title_to_find:
                # 检查是否是"第X章 标题"格式，且核心文本包含章节前缀
                chapter_pattern = re.compile(r'^第[一二三四五六七八九十百千万0-9]+章\s*(.+)$')
                match_chapter = chapter_pattern.match(title_to_find)
                if match_chapter:
                    # 提取章节后的标题部分
                    title_after_chapter = match_chapter.group(1).strip()
                    core_title_after_chapter = self._extract_core_title(title_after_chapter)
                    if core_title_after_chapter:
                        # 重新搜索，使用章节后的标题部分
                        for i in range(search_from, len(lines)):
                            if toc_start != -1 and toc_end != -1 and toc_start <= i < toc_end:
                                continue

                            line_stripped = lines[i].strip()
                            if not line_stripped:
                                continue

                            title_info = self._extract_title_from_line(line_stripped)
                            if not title_info or not title_info['is_title']:
                                continue

                            if not self._is_title_only_line(line_stripped, title_info):
                                continue

                            line_core_title = self._extract_core_title(line_stripped)

                            # 匹配章节后的标题部分
                            core1_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', core_title_after_chapter)
                            core2_clean = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', line_core_title)
                            if core1_clean == core2_clean and len(core1_clean) >= 1:
                                is_in_sub_toc = self._is_in_sub_toc(lines, i)
                                if not is_in_sub_toc:
                                    found_line = i
                                    break

            if found_line != -1:
                toc_title_positions.append((found_line, toc_header))
                last_found_line = found_line  # 更新上次找到的位置
            else:
                # 提供更详细的调试信息
                logger.warning(f"未在正文中找到目录标题: {title_to_find} (核心文本: {core_title_to_find}, 前缀: {numeric_prefix})")
                # 尝试在前100行中查找，看看是否有类似的标题
                debug_search_end = min(search_from + 100, len(lines))
                similar_titles = []
                for i in range(search_from, debug_search_end):
                    line_stripped = lines[i].strip()
                    if not line_stripped:
                        continue
                    title_info = self._extract_title_from_line(line_stripped)
                    if title_info and title_info['is_title']:
                        line_title_clean = title_info['title_clean']
                        line_core_title = self._extract_core_title(line_title_clean)
                        if line_core_title and core_title_to_find and line_core_title[:min(3, len(line_core_title))] == core_title_to_find[:min(3, len(core_title_to_find))]:
                            similar_titles.append(f"第{i+1}行: {line_stripped[:50]}")
                if similar_titles:
                    logger.debug(f"  找到相似标题: {similar_titles[:3]}")

        if not toc_title_positions:
            logger.warning("未在正文中找到任何目录中的标题")
            return []

        # 按照找到的位置排序（确保按正文中的顺序）
        toc_title_positions.sort(key=lambda x: x[0])

        # 提取章节内容：按照识别并匹配上的相邻标题之间的内容作为上一标题的文本内容
        # 注意：这里使用的是正文中实际找到的相邻标题，而不是目录中的相邻标题
        final_chunks = []
        global_counter = 1

        for idx in range(len(toc_title_positions)):
            found_line, toc_header = toc_title_positions[idx]

            # 确定章节结束位置：下一个在正文中匹配到的标题的位置，或文件末尾
            # 关键：使用正文中实际匹配到的相邻标题，而不是目录中的相邻标题
            if idx + 1 < len(toc_title_positions):
                # 下一个在正文中匹配到的标题位置
                end_line = toc_title_positions[idx + 1][0]
            else:
                # 最后一个章节：检查是否遇到结束标记（如"参考文献"）
                end_line = len(lines)
                for i in range(found_line + 1, len(lines)):
                    line_stripped = lines[i].strip()
                    if not line_stripped:
                        continue
                    end_keyword = self._check_end_section_keyword(line_stripped)
                    if end_keyword:
                        end_line = i
                        break

            # 提取章节内容（从标题行的下一行开始，到下一个标题行之前）
            if end_line <= found_line:
                continue

            if found_line + 1 >= len(lines):
                continue

            section_lines = lines[found_line + 1:end_line]
            section_content = '\n'.join(section_lines)
            section_content = clean_text_basic(section_content)

            # 不再过滤短章节，提取所有章节

            chunk_id = f"chunk_{idx+1:03d}_{global_counter:03d}"
            # 构建完整标题（包含前缀）
            numeric_prefix = toc_header.get('numeric_prefix')
            title_clean = toc_header['title']

            # 如果标题中已经包含前缀，直接使用；否则添加前缀
            # 注意：对于"第一章"、"第一节"等，前缀已经在标题中了
            if numeric_prefix:
                # 检查前缀是否已经在标题中（去除空格后比较）
                title_no_space = re.sub(r'\s+', '', title_clean)
                prefix_no_space = re.sub(r'\s+', '', str(numeric_prefix))
                if prefix_no_space in title_no_space or title_clean.startswith(numeric_prefix):
                    chunk_title = title_clean
                else:
                    chunk_title = f"{numeric_prefix} {title_clean}".strip()
            else:
                chunk_title = title_clean

            final_chunks.append({
                'chunk_id': chunk_id,
                'chunk_title': chunk_title,
                'text': section_content,
                'level': toc_header['level'],
                'parent_title': None,
                'numeric_prefix': numeric_prefix  # 保存前缀信息
            })
            global_counter += 1

        logger.info(f"严格按照目录提取到 {len(final_chunks)} 个章节")
        return final_chunks


def read_excel_paths(excel_path: str, sheet_name: str = "OCR", column_index: int = 3) -> List[str]:
    """
    从Excel文件读取指定sheet的指定列（D列，索引为3）

    Args:
        excel_path: Excel文件路径
        sheet_name: Sheet名称
        column_index: 列索引（D列为3，从0开始）

    Returns:
        路径列表
    """
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        # D列是第4列，索引为3（从0开始）
        paths = df.iloc[:, column_index].dropna().astype(str).tolist()
        # 过滤空字符串
        paths = [p.strip() for p in paths if p.strip() and p.strip().lower() != 'nan']
        return paths
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        raise


def find_file_path(file_path_in_excel: str) -> Optional[str]:
    """
    根据Excel中的文件路径，找到实际文件

    例如：./data/books/OCR/Agri/20260108/9787040470406.md
    则去 ./data/books/OCR/Agri/20260108/ 下找 9787040470406.md
    """
    if not file_path_in_excel or not file_path_in_excel.strip():
        return None

    file_path_in_excel = file_path_in_excel.strip()

    # 如果路径已经是完整路径且文件存在，直接返回
    if os.path.exists(file_path_in_excel):
        return file_path_in_excel

    # 提取目录和文件名
    # 例如：./data/books/OCR/Agri/20260108/9787040470406.md
    # 目录：./data/books/OCR/Agri/20260108/
    # 文件名：9787040470406.md
    dir_path = os.path.dirname(file_path_in_excel)
    file_name = os.path.basename(file_path_in_excel)

    # 在目录下查找文件
    if dir_path and os.path.isdir(dir_path):
        full_path = os.path.join(dir_path, file_name)
        if os.path.exists(full_path):
            return full_path

    # 如果找不到，尝试直接使用原路径
    logger.warning(f"未找到文件: {file_path_in_excel}")
    return None


# 已删除章节统计功能（process_file函数）
# 现在脚本只用于QA生成，直接从JSON文件读取数据



    def _title_similar(self, title1: str, title2: str, threshold: float = 0.6) -> bool:
        """
        判断两个标题是否相似（改进的字符串相似度算法）
        使用字符级别的匹配，要求更严格
        """
        if not title1 or not title2:
            return False
        # 去除空格和标点后比较
        t1_clean = re.sub(r'[\s\W]+', '', title1)
        t2_clean = re.sub(r'[\s\W]+', '', title2)
        if not t1_clean or not t2_clean:
            return False
        # 如果完全相等，直接返回True
        if t1_clean == t2_clean:
            return True

        # 使用更严格的相似度计算：计算相同字符的比例
        # 将字符串转换为字符集合，计算交集和并集
        set1 = set(t1_clean)
        set2 = set(t2_clean)

        # 计算交集（共同字符）
        intersection = set1 & set2
        # 计算并集（所有字符）
        union = set1 | set2

        if len(union) == 0:
            return False

        # Jaccard相似度：交集大小 / 并集大小
        jaccard_similarity = len(intersection) / len(union) if len(union) > 0 else 0.0

        # 同时检查字符顺序的相似度（简单的字符匹配）
        # 计算相同位置的字符匹配数
        min_len = min(len(t1_clean), len(t2_clean))
        max_len = max(len(t1_clean), len(t2_clean))
        if max_len == 0:
            return False

        # 计算相同位置的字符匹配数
        match_count = sum(1 for i in range(min_len) if t1_clean[i] == t2_clean[i])
        position_similarity = match_count / max_len if max_len > 0 else 0.0

        # 综合相似度：取Jaccard相似度和位置相似度的平均值
        # 但要求两者都达到一定阈值
        combined_similarity = (jaccard_similarity + position_similarity) / 2.0

        # 对于高阈值（0.9），要求更严格：Jaccard和位置相似度都要高
        if threshold >= 0.9:
            return jaccard_similarity >= 0.85 and position_similarity >= 0.85

        return combined_similarity >= threshold

    def _match_title_strict(self, target_title: str, target_numeric: Optional[str],
                           line_text: str, line_numeric: Optional[str]) -> bool:
        """
        严格匹配标题：要求序号和标题内容同时匹配
        忽略空格和页码，但序号和标题内容都必须匹配
        """
        # 如果都有数字前缀，序号和标题内容都必须匹配
        if target_numeric and line_numeric:
            # 步骤1：数字前缀（序号）必须完全匹配（忽略空格）
            target_num_clean = re.sub(r'\s+', '', target_numeric)
            line_num_clean = re.sub(r'\s+', '', line_numeric)
            if target_num_clean != line_num_clean:
                # 序号不匹配，直接返回False
                return False

            # 步骤2：序号匹配后，检查标题文本（去除数字前缀、空格和页码）
            target_text = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s*', '', target_title).strip()
            line_text_clean = re.sub(r'^[0-9]+(?:\.[0-9]+)*\s*', '', line_text).strip()
            # 去除页码（行末的数字）
            target_text = re.sub(r'\s+[0-9]+\s*$', '', target_text).strip()
            line_text_clean = re.sub(r'\s+[0-9]+\s*$', '', line_text_clean).strip()

            # 步骤3：标题内容必须匹配（对于有序号的标题，要求完全匹配，不使用相似度匹配）
            target_text_no_space = re.sub(r'\s+', '', target_text)
            line_text_no_space = re.sub(r'\s+', '', line_text_clean)

            # 对于有序号的标题，要求完全匹配（去除空格后）
            # 不使用相似度匹配，避免误匹配
            if target_text_no_space == line_text_no_space:
                return True

            # 如果去除空格后不完全相等，返回False（有序号的标题必须完全匹配）
            return False
        else:
            # 没有数字前缀，只匹配标题文本（忽略空格和页码）
            target_clean = re.sub(r'\s+[0-9]+\s*$', '', target_title).strip()
            line_clean = re.sub(r'\s+[0-9]+\s*$', '', line_text).strip()
            # 去除所有空格后比较
            target_no_space = re.sub(r'\s+', '', target_clean)
            line_no_space = re.sub(r'\s+', '', line_clean)
            if target_no_space == line_no_space:
                return True
            # 使用高阈值相似度匹配（阈值0.9，更严格）
            return self._title_similar(target_clean, line_clean, threshold=0.9)

    def _split_by_specific_level(self, text: str, level: int, parent_start_line: int = 0) -> List[Dict[str, str]]:
        """按指定级别分割文本"""
        lines = text.split('\n')
        chunks = []
        current_chunk = []
        current_title = f"Level_{level}_Root"
        current_start_line = parent_start_line
        in_chunk = False

        for i, line in enumerate(lines):
            header_match = self.markdown_header_pattern.match(line.strip())

            if header_match and len(header_match.group(1)) == level:
                if in_chunk and current_chunk:
                    chunk_content = clean_text_basic('\n'.join(current_chunk))
                    if len(chunk_content) >= MIN_CHAPTER_LENGTH:
                        chunks.append({
                            'start_line': current_start_line,
                            'title': current_title,
                            'level': level,
                            'content': chunk_content,
                            'raw_content': '\n'.join(current_chunk)
                        })

                current_title = header_match.group(2).strip()
                current_chunk = [line]
                current_start_line = parent_start_line + i
                in_chunk = True

            elif in_chunk:
                current_chunk.append(line)
            elif not in_chunk and header_match and len(header_match.group(1)) < level:
                break

        if in_chunk and current_chunk:
            chunk_content = clean_text_basic('\n'.join(current_chunk))
            if len(chunk_content) >= MIN_CHAPTER_LENGTH:
                chunks.append({
                    'start_line': current_start_line,
                    'title': current_title,
                    'level': level,
                    'content': chunk_content,
                    'raw_content': '\n'.join(current_chunk)
                })

        return chunks

    def _fallback_split(self, text: str) -> List[Dict[str, str]]:
        """回退分块策略"""
        content = clean_text_basic(text)
        if len(content) < MIN_CHAPTER_LENGTH:
            logger.warning(f"文本内容过短 ({len(content)} < {MIN_CHAPTER_LENGTH})，跳过")
            return []

        num_chunks = min(MAX_CHUNK_SIZE, max(1, (len(content) + IDEA_CHUNK_LENGTH - 1) // IDEA_CHUNK_LENGTH))
        chunk_size = len(content) // num_chunks

        chunks = []
        for i in range(num_chunks):
            start = i * chunk_size
            end = len(content) if i == num_chunks - 1 else start + chunk_size
            chunk_text = content[start:end]
            chunk_id = f"chunk_001_01_{i+1:03d}"
            chunks.append({
                'chunk_id': chunk_id,
                'chunk_title': f'块 {i+1}',
                'text': chunk_text,
                'level': 1,
                'parent_title': None
            })
        logger.info(f"未检测到标题结构，按长度分割为 {num_chunks} 个chunks")
        return chunks

    def _merge_final_chunks(self, chunks: List[Dict], max_chunks: int) -> List[Dict]:
        """合并最终块以不超过限制"""
        if len(chunks) <= max_chunks:
            return chunks

        merge_ratio = len(chunks) // max_chunks + 1
        merged_chunks = []
        global_counter = 1

        for i in range(0, len(chunks), merge_ratio):
            group_chunks = chunks[i:i + merge_ratio]
            merged_content = ' '.join(chunk['text'] for chunk in group_chunks)

            if len(group_chunks) == 1:
                merged_title = group_chunks[0]['chunk_title']
            else:
                first_title = group_chunks[0]['chunk_title']
                last_title = group_chunks[-1]['chunk_title']
                merged_title = f"{first_title} ~ {last_title}"

            first_chunk_id = group_chunks[0]['chunk_id']
            id_parts = first_chunk_id.split('_')
            l2_id = id_parts[1] if len(id_parts) > 1 else "001"

            merged_chunks.append({
                'chunk_id': f"chunk_{l2_id}_01_{global_counter:03d}",
                'chunk_title': merged_title,
                'text': merged_content,
                'level': 2,
                'parent_title': None
            })
            global_counter += 1

        return merged_chunks

# --- SFT问答生成器 ---
class SFTQuestionGenerator:
    def __init__(self):
        self.client = OpenAI(api_key=OPENAI_API_KEY, base_url=OPENAI_BASE_URL, timeout=180.0)

    def build_prompt(self, chunk_text: str, chunk_title: str, max_q: int = 5) -> str:
        """构建生成SFT问答对的提示词"""
        forbidden_examples = "\n".join([f"   - '{phrase}'" for phrase in FORBIDDEN_PHRASES[:15]])

        system_prompt = f"""你是一位农业与生命科学领域的专业问答生成系统，为科研人员和学术工作者提供准确、可验证的专业知识。

【输出要求】
- 所有内容必须严格基于给定文本，不得加入外部知识或推测
- 问题应具有通用性，不依赖具体研究背景
- 答案应描述一般规律而非具体数据
- 严格输出JSON数组格式，无额外说明

【严格禁止】
1. 问题和答案中严格禁止使用以下指代图书的表述：
{forbidden_examples}
2. 问题中禁止围绕以下主题发问：
   - '本书内容'、'该章节'、'该部分'
   - 针对具体的图和表
   - 具体的数值、参数等细节
3. 所有问题必须在脱离本文（例如换成同一领域的任意一本类似书籍）时仍然合理成立

【质量要求】
- 问题通用性
- 答案一般化
- 语言规范性
- 术语准确性

【输出格式】
[
  {{
    "question": "问题内容",
    "answer": "科学事实、机制或原理",
    "difficulty": "easy | medium | hard",
    "tags": ["tag1", "tag2"]
  }}
]"""

        # 对单次输入正文做长度截断，默认允许最多约30万字符
        processed_text = (
            chunk_text[:300000] + "\n\n[以下内容因长度被截断]"
            if len(chunk_text) > 300000
            else chunk_text
        )

        user_prompt = f"""任务：为以下图书章节生成 {max_q} 组高质量问答对。

【章节信息】
名称：{chunk_title}
内容：
{processed_text}

【核心要求】
1. 有效性检查
2. 零幻觉原则：所有答案内容必须严格基于给定文本，不加入外部知识
3. 内容聚焦：核心科学概念、研究方法、机制原理
4. 问题设计：关注通用概念
5. 元数据：难度分级 + 精确标签

请严格遵循JSON格式输出。"""

        return f"{system_prompt}\n\n{user_prompt}"

    def generate_for_chunk(
        self,
        chunk: Dict[str, str],
        source_id: str,
        enable_quality_filter: bool = False,
        min_quality_score: float = 60.0
    ) -> List[Dict]:
        """为单个chunk生成SFT问答对（简单生成模式）"""
        prompt = self.build_prompt(chunk['text'], chunk['chunk_title'], MAX_Q_PER_CHUNK)
        start_time = time.time()

        try:
            logger.info(f"开始为来源 {source_id} 的章节 '{chunk['chunk_id']}-{chunk['chunk_title']}' 生成问答对")

            # 增加LLM调用计数
            call_count = increment_llm_call_count()
            print(f"  \033[1;36;40m[LLM调用 #{call_count}]\033[0m 正在调用模型: {DEFAULT_MODEL}")

            response = self.client.chat.completions.create(
                model=DEFAULT_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7,
                max_tokens=8000,
                stream=False
            )
            latency = time.time() - start_time
            content = response.choices[0].message.content.strip()

            usage = response.usage
            input_tokens = getattr(usage, 'prompt_tokens', 0)
            output_tokens = getattr(usage, 'completion_tokens', 0)

            logger.info(f"[{source_id}] API调用成功 - 输入Tokens: {input_tokens}, 输出Tokens: {output_tokens}, 耗时: {latency:.2f}s")

            # 提取JSON
            json_str = content
            m = re.search(r'\[\s*{', content)
            if m:
                start = m.start()
                end = content.rfind(']')
                if end != -1 and end > start:
                    json_str = content[start:end+1]

            try:
                qas_data = json.loads(json_str)
            except Exception as e:
                logger.error(f"解析 JSON 失败，section={chunk['chunk_title']}, error={e}")
                return []

            if not isinstance(qas_data, list):
                logger.warning(f"返回数据不是列表格式，section={chunk['chunk_title']}")
                return []

            qas = []
            quality_scorer = QualityScorer()  # 移到循环外，避免重复创建
            for item in qas_data:
                if not isinstance(item, dict):
                    continue

                q = str(item.get("question", "")).strip()
                a = str(item.get("answer", "")).strip()
                difficulty = str(item.get("difficulty", "")).strip().lower()
                tags = item.get("tags", [])

                if not q or not a or len(q) < 8 or len(a) < 20:
                    continue

                # 清洗文本
                q = re.sub(r'\n+', ' ', q)
                a = re.sub(r'\n+', ' ', a)
                q = re.sub(r'\s{2,}', ' ', q).strip()
                a = re.sub(r'\s{2,}', ' ', a).strip()

                # 质量检查
                if is_study_dependent(q) or is_study_dependent(a):
                    continue

                if any(phrase in q for phrase in FORBIDDEN_PHRASES) or \
                   any(phrase in a for phrase in FORBIDDEN_PHRASES):
                    continue

                q_clean = sanitize_text_forbidden_phrases(q)
                a_clean = sanitize_text_forbidden_phrases(a)

                if not q_clean or not a_clean:
                    continue

                bad_pattern = r'(文中|本文|文章中|根据文本|根据以上文本|根据上述文本|给定文本|该章节|该部分)'
                if re.search(bad_pattern, q_clean) or re.search(bad_pattern, a_clean):
                    continue

                if difficulty not in ["easy", "medium", "hard"]:
                    if len(q_clean) < 40:
                        difficulty = "easy"
                    elif len(q_clean) < 80:
                        difficulty = "medium"
                    else:
                        difficulty = "hard"

                if not isinstance(tags, list):
                    tags = [str(tags)]

                # 构建SFT格式的记录
                record = {
                    "source_id": source_id,
                    "paper_id": source_id,  # 添加paper_id字段
                    "source_type": "book",
                    "chapter_title": chunk['chunk_title'],  # 添加chapter_title字段
                    "question": q_clean,
                    "answer": a_clean,
                    "reasoning_steps": [],
                    "question_cot": "",
                    "final_conclusion": "",
                    "difficulty": difficulty,
                    "curriculum_stage": assign_curriculum_stage(difficulty, ""),
                    "tags": tags,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "generation_type": "简单型",
                }

                # 质量评分
                record["quality_report"] = quality_scorer.score_qa_pair(record)

                qas.append(record)

            # 质量过滤（如果启用）
            if enable_quality_filter:
                # 复用上面创建的 quality_scorer 实例
                qas = quality_scorer.filter_by_quality(qas, min_quality_score)

            logger.info(f"成功为章节 '{chunk['chunk_id']}-{chunk['chunk_title']}' 生成 {len(qas)} 个问答对")
            return qas

        except Exception as e:
            logger.error(f"章节 '{chunk['chunk_title']}' 生成失败: {e}")
            return []

    def _generate_simple_qa_fallback(
        self,
        section_name: str,
        section_text: str,
        max_q: int = 5,
    ) -> List[Dict[str, Any]]:
        """
        降级策略：当推理链抽取失败时，直接生成简单QA

        Args:
            section_name: 章节名称
            section_text: 章节文本
            max_q: 最大问题数

        Returns:
            List[Dict[str, Any]]: QA列表
        """
        try:
            # 使用简单的prompt直接生成QA
            simple_prompt = f"""你是一位农业与生命科学领域的专业问答生成系统。

请基于以下章节内容，生成 {max_q} 个问答对。

章节名称：{section_name}
章节内容：
\"\"\"markdown
{section_text[:15000]}  # 限制长度避免过长
\"\"\"

【要求】
1. 问题应具有通用性，不依赖具体研究背景
2. 答案应描述一般规律而非具体数据
3. 严格输出JSON数组格式：
[
  {{
    "question": "问题文本",
    "answer": "答案文本",
    "difficulty": "easy|medium|hard",
    "tags": ["tag1", "tag2"]
  }}
]

【禁止】
- 不要使用"该章节"、"本文"、"书中"等指代表述
- 不要引用具体数值、浓度、时间等细节
"""

            qa_data = call_responses_for_json(
                prompt=simple_prompt,
                model=DEFAULT_MODEL,
                max_output_tokens=8000,
            )

            # 处理返回的QA数据
            if isinstance(qa_data, list):
                qa_list = qa_data
            elif isinstance(qa_data, dict):
                if "qas" in qa_data:
                    qa_list = qa_data.get("qas", [])
                elif "questions" in qa_data:
                    qa_list = qa_data.get("questions", [])
                else:
                    qa_list = [qa_data]
            else:
                qa_list = []

            # 转换为标准格式
            result = []
            for qa in qa_list[:max_q]:
                q = str(qa.get("question", "")).strip()
                a = str(qa.get("answer", "")).strip()
                difficulty = str(qa.get("difficulty", "medium")).strip().lower()
                tags = qa.get("tags", [])

                if not isinstance(tags, list):
                    tags = [str(tags)]

                if q and a and len(q) >= 8 and len(a) >= 20:
                    result.append({
                        "question": q,
                        "answer": a,
                        "reasoning_steps": [],
                        "question_cot": "",
                        "final_conclusion": "",
                        "difficulty": difficulty if difficulty in ["easy", "medium", "hard"] else "medium",
                        "tags": tags,
                    })

            if result:
                logger.info(f"降级策略成功生成 {len(result)} 个简单QA")
                return result
            else:
                logger.warning(f"降级策略未能生成有效QA")
                return []

        except Exception as e:
            logger.warning(f"降级策略失败: {e}")
            return []

    def generate_reasoning_qas_from_section(
        self,
        section_name: str,
        section_text: str,
        max_q: int = 5,
    ) -> Tuple[List[Dict[str, Any]], str]:
        """
        两阶段流水线：
        1) 从 section 文本抽取多条 reasoning chains
        2) 每条 chain 生成一题需要多步推理的 QA（带 cot）
        3) 对生成的 QA 做清洗/过滤，返回统一结构

        Args:
            section_name: 章节名称
            section_text: 章节文本
            max_q: 最大问题数

        Returns:
            Tuple[List[Dict[str, Any]], str]: (qas列表, generation_type)
        """
        generation_type = "推理型"

        # 1) 抽取推理链
        chain_prompt = build_chain_extraction_prompt(
            section_name=section_name,
            section_text=section_text,
            max_chains=max_q
        )
        try:
            chain_data = call_responses_for_json(
                prompt=chain_prompt,
                model=DEFAULT_MODEL,
                max_output_tokens=8000,
            )
        except Exception as e:
            # 只在debug模式下显示详细错误，避免输出过多
            logger.debug(f"推理链抽取失败，section={section_name}, 错误: {e}")
            tprint(f"  ❌ 推理链抽取失败，section={section_name}，尝试降级到简单QA生成")
            # 降级策略：如果推理链抽取失败，尝试直接生成简单QA
            return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"

        chains = []
        if isinstance(chain_data, dict) and "chains" in chain_data:
            chains = chain_data.get("chains", [])
        elif isinstance(chain_data, list):
            chains = chain_data
        else:
            # 尝试从非标准格式中提取信息
            tprint(f"  ⚠️ 推理链返回结构异常，section={section_name}，尝试提取部分信息")
            if isinstance(chain_data, dict):
                # 尝试提取其他可能的字段
                if "data" in chain_data:
                    chains = chain_data.get("data", [])
                elif "result" in chain_data:
                    chains = chain_data.get("result", [])
                else:
                    # 如果提取失败，尝试降级
                    logger.debug(f"无法从异常结构中提取chains，尝试降级")
                    return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"
            elif isinstance(chain_data, str):
                # 如果是字符串，尝试解析为JSON
                try:
                    parsed = json.loads(chain_data)
                    if isinstance(parsed, dict) and "chains" in parsed:
                        chains = parsed.get("chains", [])
                    elif isinstance(parsed, list):
                        chains = parsed
                    else:
                        return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"
                except Exception:
                    return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"
            else:
                # 其他类型，尝试降级
                return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"

        if not isinstance(chains, list) or not chains:
            tprint(f"  ⚠️ 未抽取到有效推理链，section={section_name}，尝试降级到简单QA生成")
            # 降级策略：如果未抽取到推理链，尝试直接生成简单QA
            return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"

        # 2) 每条 chain 生成一题 QA
        raw_qas = []
        # 控制最多生成 max_q 题
        for chain in chains:
            if len(raw_qas) >= max_q:
                break
            try:
                # 保存第一阶段推理链的steps（来自图书的推理逻辑）
                reasoning_steps = chain.get("steps", [])
                final_conclusion = chain.get("final_conclusion", "")

                chain_json_str = json.dumps(chain, ensure_ascii=False)
                qa_prompt = build_chain_to_qa_prompt(chain_json_str)
                qa_data = call_responses_for_json(
                    prompt=qa_prompt,
                    model=DEFAULT_MODEL,
                    max_output_tokens=8000,
                )
                if isinstance(qa_data, list):
                    qa_list = qa_data
                else:
                    qa_list = [qa_data]

                for qa in qa_list:
                    if len(raw_qas) >= max_q:
                        break
                    q = str(qa.get("question", "")).strip()
                    a = str(qa.get("answer", "")).strip()
                    meta = qa.get("meta", {}) or {}
                    difficulty = str(meta.get("difficulty", qa.get("difficulty", ""))).strip().lower()
                    tags = meta.get("tags", qa.get("tags", [])) or []
                    cot_raw = qa.get("cot", "")

                    # 标准化第二阶段cot：可能是 list 或 str
                    if isinstance(cot_raw, list):
                        question_cot = "\n".join(str(s).strip() for s in cot_raw if str(s).strip())
                    else:
                        question_cot = str(cot_raw or "").strip()

                    raw_qas.append({
                        "question": q,
                        "answer": a,
                        "reasoning_steps": reasoning_steps,  # 第一阶段：图书推理链steps
                        "question_cot": question_cot,        # 第二阶段：针对问题的推理链
                        "difficulty": difficulty,
                        "tags": tags,
                        "final_conclusion": final_conclusion,  # 推理链的结论
                    })
            except Exception as e:
                tprint(f"  ⚠️ 从推理链生成 QA 失败: {e}")
                continue

        if not raw_qas:
            tprint(f"  ⚠️ 推理链生成 QA 为空，section={section_name}，尝试降级到简单QA生成")
            # 降级策略：如果推理链生成QA失败，尝试直接生成简单QA
            return self._generate_simple_qa_fallback(section_name, section_text, max_q), "简单型（降级）"

        # 3) 清洗/过滤，与原逻辑保持一致风格
        qas: List[Dict[str, Any]] = []
        for item in raw_qas:
            q = str(item.get("question", "")).strip()
            a = str(item.get("answer", "")).strip()
            difficulty = str(item.get("difficulty", "")).strip().lower()
            tags = item.get("tags", [])
            reasoning_steps = item.get("reasoning_steps", [])
            question_cot = str(item.get("question_cot", "")).strip()
            final_conclusion = str(item.get("final_conclusion", "")).strip()

            if not isinstance(tags, list):
                tags = [str(tags)]

            if not q or not a or len(q) < 8 or len(a) < 20:
                continue

            q = re.sub(r'\n+', ' ', q)
            a = re.sub(r'\n+', ' ', a)
            q = re.sub(r'\s{2,}', ' ', q).strip()
            a = re.sub(r'\s{2,}', ' ', a).strip()

            if is_study_dependent(q) or is_study_dependent(a):
                continue

            if any(phrase in q for phrase in FORBIDDEN_PHRASES) or \
               any(phrase in a for phrase in FORBIDDEN_PHRASES):
                continue

            q_clean = sanitize_text_forbidden_phrases(q)
            a_clean = sanitize_text_forbidden_phrases(a)

            if not q_clean or not a_clean:
                continue

            bad_pattern = r'(文中|本文|文章中|根据文本|根据以上文本|根据上述文本|给定文本|该章节|该部分)'
            if re.search(bad_pattern, q_clean) or re.search(bad_pattern, a_clean):
                continue

            if difficulty not in ["easy", "medium", "hard"]:
                if len(q_clean) < 40:
                    difficulty = "easy"
                elif len(q_clean) < 80:
                    difficulty = "medium"
                else:
                    difficulty = "hard"

            # 清洗两个阶段的CoT
            question_cot_clean = sanitize_text_forbidden_phrases(question_cot) if question_cot else ""
            final_conclusion_clean = sanitize_text_forbidden_phrases(final_conclusion) if final_conclusion else ""

            qas.append({
                "question": q_clean,
                "answer": a_clean,
                "reasoning_steps": reasoning_steps,      # 第一阶段：图书推理链steps
                "question_cot": question_cot_clean,      # 第二阶段：针对问题的推理链
                "final_conclusion": final_conclusion_clean,  # 推理链的结论
                "difficulty": difficulty,
                "tags": tags,
            })

        return qas, generation_type

    def generate_for_chunk_with_reasoning(
        self,
        chunk: Dict[str, str],
        source_id: str,
        enable_diversity: bool = True,
        simhash_dedup_hamming: int = 6,
        enable_quality_filter: bool = False,
        min_quality_score: float = 60.0
    ) -> List[Dict]:
        """为单个chunk生成SFT问答对（支持推理链，集成质量评估和统计）"""
        # 始终使用推理链生成
        tprint(f"▶ 使用推理链生成: {chunk['chunk_title']}")
        raw_qas, generation_type = self.generate_reasoning_qas_from_section(
            section_name=chunk['chunk_title'],
            section_text=chunk['text'],
            max_q=MAX_Q_PER_CHUNK * OVER_GENERATE_FACTOR,
        )

        if not raw_qas:
            logger.debug(f"章节 '{chunk.get('chunk_title', 'unknown')}' 未生成有效问答对")
            # 只在debug模式下显示，避免输出过多
            # tprint("  - 未生成有效问答对")
            return []

        # 构建完整的SFT格式记录
        records = []
        raw_count = len(raw_qas)
        invalid_count = 0
        valid_count = 0

        quality_scorer = QualityScorer()

        for qa in raw_qas:
            record = {
                "source_id": source_id,
                "paper_id": source_id,  # 添加paper_id字段
                "source_type": "book",
                "chapter_title": chunk['chunk_title'],  # 添加chapter_title字段
                "question": qa["question"],
                "answer": qa["answer"],
                "reasoning_steps": qa.get("reasoning_steps", []),
                "question_cot": qa.get("question_cot", ""),
                "final_conclusion": qa.get("final_conclusion", ""),
                "difficulty": qa.get("difficulty", "medium"),
                "curriculum_stage": assign_curriculum_stage(
                    qa.get("difficulty", "medium"),
                    qa.get("question_cot", "")
                ),
                "tags": qa.get("tags", []),
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "generation_type": generation_type,
            }

            # 质量评分
            record["quality_report"] = quality_scorer.score_qa_pair(record)

            # 基本验证
            if len(record["question"]) >= 8 and len(record["answer"]) >= 20:
                records.append(record)
                valid_count += 1
            else:
                invalid_count += 1

        if not records:
            tprint(f"  ⚠️ {generation_type} raw={raw_count} → valid=0 (all invalid)")
            return []

        # 先按质量降序排序
        records.sort(
            key=lambda x: (x.get("quality_report", {}).get("total_score", 0.0)),
            reverse=True
        )

        # SimHash 去重统计
        simhash_before = len(records)
        records = dedup_qas_simhash(records, max_hamming=simhash_dedup_hamming)
        simhash_after = len(records)

        # 多样性过滤统计
        diversity_before = len(records)
        if enable_diversity and (diversity_filter_qas is not None):
            records = diversity_filter_qas(records, max_keep=MAX_Q_PER_CHUNK * OVER_GENERATE_FACTOR, jaccard_threshold=0.7)
        diversity_after = len(records)

        # 质量过滤
        quality_before = len(records)
        if enable_quality_filter:
            records = quality_scorer.filter_by_quality(records, min_quality_score)
        quality_after = len(records)

        # 最终截断统计
        final_before_trunc = len(records)
        records = records[:MAX_Q_PER_CHUNK]
        final_after_trunc = len(records)

        # 质量统计
        stats = quality_scorer.get_quality_statistics(records)

        # 将 chunk 统计写回每条 record
        chunk_stats = {
            "raw_count": raw_count,
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "simhash_before": simhash_before,
            "simhash_after": simhash_after,
            "diversity_before": diversity_before,
            "diversity_after": diversity_after,
            "quality_before": quality_before,
            "quality_after": quality_after,
            "final_before_trunc": final_before_trunc,
            "final_after_trunc": final_after_trunc,
            "avg_quality_score": stats.get("average_score", 0),
            "pass_rate": stats.get("pass_rate", 0),
            "chunk_id": chunk.get("chunk_id"),
            "chunk_title": chunk.get("chunk_title"),
        }

        for r in records:
            r["chunk_stats"] = chunk_stats
            # 平铺关键字段（方便后续直接 sum/print）
            r["raw_count"] = raw_count
            r["valid_count"] = valid_count
            r["simhash_before"] = simhash_before
            r["simhash_after"] = simhash_after
            r["diversity_before"] = diversity_before
            r["diversity_after"] = diversity_after
            r["quality_before"] = quality_before
            r["quality_after"] = quality_after
            r["final_before_trunc"] = final_before_trunc
            r["final_after_trunc"] = final_after_trunc

        tprint(
            f"  ✅ {generation_type} "
            f"raw={raw_count} → valid={valid_count} "
            f"| simhash={simhash_before}->{simhash_after} "
            f"| diversity={diversity_before}->{diversity_after} "
            f"| quality={quality_before}->{quality_after} "
            f"| keep={final_after_trunc} "
            f"| avg={stats.get('average_score', 0):.1f}"
        )
        return records

# --- I/O 函数 ---
def read_excel_paths(excel_path: str, sheet_name: str = 'OCR', column_index: int = 3) -> List[str]:
    """
    从Excel文件中读取指定sheet的指定列路径

    Args:
        excel_path: Excel文件路径
        sheet_name: Sheet名称（默认'OCR'）
        column_index: 列索引，从0开始（默认3，即D列）

    Returns:
        List[str]: 路径列表
    """
    paths = []
    try:
        if HAS_PANDAS:
            # 使用pandas读取，尝试不同的sheet名称格式
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            except ValueError:
                # 如果指定的sheet名称不存在，尝试其他格式
                sheet_name_variants = [sheet_name, sheet_name.lower(), sheet_name.upper(),
                                      'OCR', 'ocr', 'Ocr']
                df = None
                for variant in sheet_name_variants:
                    try:
                        df = pd.read_excel(excel_path, sheet_name=variant, header=None)
                        break
                    except ValueError:
                        continue
                if df is None:
                    raise ValueError(f"无法找到工作表: {sheet_name} 或其变体")
            # 读取指定列，跳过空值
            column_data = df.iloc[:, column_index]
            paths = [str(path).strip() for path in column_data if pd.notna(path) and str(path).strip()]
        elif HAS_OPENPYXL:
            # 使用openpyxl读取
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                # 尝试不同的sheet名称格式
                sheet_name_variants = [sheet_name, sheet_name.lower(), sheet_name.upper(),
                                      'OCR', 'ocr', 'Ocr']
                for variant in sheet_name_variants:
                    if variant in wb.sheetnames:
                        sheet_name = variant
                        break
                else:
                    logger.error(f"未找到sheet: {sheet_name}，可用sheets: {wb.sheetnames}")
                    return []

            ws = wb[sheet_name]
            # 读取指定列（column_index+1因为Excel列从1开始）
            for row in ws.iter_rows(min_row=1, min_col=column_index+1, max_col=column_index+1, values_only=True):
                if row[0] and str(row[0]).strip():
                    paths.append(str(row[0]).strip())
        else:
            logger.error("需要安装pandas或openpyxl来读取Excel文件")
            return []

        logger.info(f"从Excel文件读取到 {len(paths)} 个路径")
        return paths
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        return []

def find_file_in_directory(file_path: str) -> Optional[str]:
    """
    根据完整路径，在文件所在目录中查找文件

    Args:
        file_path: 完整文件路径，如 ./data/books/OCR/Agri/20260105/9787040308198.md

    Returns:
        Optional[str]: 找到的文件路径，如果不存在则返回None
    """
    if os.path.isfile(file_path):
        return file_path

    # 提取目录和文件名
    file_dir = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)

    # 在目录中查找文件
    if os.path.isdir(file_dir):
        full_path = os.path.join(file_dir, file_name)
        if os.path.isfile(full_path):
            return full_path
        else:
            logger.warning(f"文件不存在: {full_path}")
            return None
    else:
        logger.warning(f"目录不存在: {file_dir}")
        return None

def read_jsonl(path: str) -> List[Dict]:
    """读取JSONL文件（单文件），支持新格式：books_ID, chapter_title, context, length"""
    logger.info(f"开始读取输入文件(单文件): {path}")
    chapters = []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    item = json.loads(line)
                    # 支持新格式：books_ID, chapter_title, context, length
                    book_id = item.get('books_ID', item.get('id', 'unknown'))
                    chapter_title = item.get('chapter_title', '')
                    context = item.get('context', item.get('text', ''))
                    # 确保length是整数类型
                    try:
                        length = int(item.get('length', 0))
                    except (ValueError, TypeError):
                        length = 0

                    # 如果length为0或小于100，跳过（根据length判断是否生成QA）
                    if length == 0 or length < 100:
                        continue

                    chapters.append({
                        'books_ID': book_id,
                        'chapter_title': chapter_title,
                        'context': context,
                        'length': length,
                    })
        logger.info(f"成功读取 {len(chapters)} 个章节（根据 books_ID/chapter_title/context/length 字段）")
    except FileNotFoundError:
        logger.error(f"文件未找到: {path}")
        raise
    except Exception as e:
        logger.error(f"读取文件失败: {e}")
        raise
    return chapters


def read_input(path: str) -> List[Dict]:
    """
    读取输入路径：
    - 如果是文件：
      - .json/.jsonl 文件：按 JSONL 读取（新格式：books_ID, chapter_title, context, length）
    - 如果是目录：读取目录下所有 .json/.jsonl 文件并合并
    """
    if os.path.isdir(path):
        logger.info(f"输入为目录，将读取其中的JSON/JSONL文件: {path}")
        all_chapters: List[Dict[str, Any]] = []
        for fname in sorted(os.listdir(path)):
            if not (fname.endswith('.json') or fname.endswith('.jsonl')):
                continue
            fpath = os.path.join(path, fname)
            logger.info(f"开始读取子文件: {fpath}")
            sub_chapters = read_jsonl(fpath)
            logger.info(f"子文件 {fname} 读取到 {len(sub_chapters)} 个章节")
            all_chapters.extend(sub_chapters)
        logger.info(f"目录 {path} 合计读取 {len(all_chapters)} 个章节")
        return all_chapters
    else:
        # 默认按单文件JSONL处理
        return read_jsonl(path)

def save_qas(qas: List[Dict], path: str, append: bool = False):
    """保存问答对到文件

    Args:
        qas: 问答对列表
        path: 文件路径
        append: 是否追加模式（默认False，覆盖写入）
    """
    try:
        # 如果路径包含目录，创建目录
        dir_path = os.path.dirname(path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        mode = 'a' if append else 'w'
        with open(path, mode, encoding='utf-8') as f:
            for qa in qas:
                f.write(json.dumps(qa, ensure_ascii=False) + '\n')
        action = "追加" if append else "保存"
        logger.info(f"成功{action} {len(qas)} 个问答对至: {path}")
    except Exception as e:
        logger.error(f"保存文件失败: {path}, 错误: {e}")
        raise

def count_qas_in_file(file_path: str) -> int:
    """统计文件中已保存的QA数量"""
    try:
        if not os.path.exists(file_path):
            return 0
        count = 0
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    count += 1
        return count
    except Exception:
        return 0

def get_processed_chapters(file_path: str) -> set:
    """
    从输出文件中读取已处理的章节名称（chapter_title字段）

    Args:
        file_path: 输出文件路径

    Returns:
        set: 已处理的章节名称集合
    """
    processed_chapters = set()
    try:
        if not os.path.exists(file_path):
            return processed_chapters

        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    qa = json.loads(line)
                    chapter_title = qa.get('chapter_title')
                    if chapter_title:
                        processed_chapters.add(chapter_title)
                except (json.JSONDecodeError, KeyError):
                    continue
    except Exception as e:
        logger.warning(f"读取已处理章节失败: {e}")

    return processed_chapters

def save_qa_single(qa: Dict, path: str):
    """实时保存单个问答对到文件（追加模式）"""
    try:
        # 如果路径包含目录，创建目录
        dir_path = os.path.dirname(path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        with open(path, 'a', encoding='utf-8') as f:
            f.write(json.dumps(qa, ensure_ascii=False) + '\n')
    except Exception as e:
        logger.error(f"实时保存问答对失败: {path}, 错误: {e}")
        raise

# --- 线程处理函数 ---
def process_single_file(
    input_path: str,
    thread_id: int,
    args: argparse.Namespace
) -> Tuple[str, List[Dict], List[Dict]]:
    """
    处理单个文件（用于多线程）

    Returns:
        Tuple[output_path, qas_list, stats_list]
    """
    # 提取目录名和文件名，格式：20260105/9787040487039.json（在try外部定义，确保异常处理可用）
    dir_name = os.path.basename(os.path.dirname(input_path))
    file_name = os.path.basename(input_path)
    file_display = f"{dir_name}/{file_name}" if dir_name else file_name

    try:
        # 更新状态：开始处理
        status_manager.update_status(thread_id, file_display, 0, "开始处理")

        # 读取当前文件的章节数据（新格式：books_ID, chapter_title, context, length）
        chapters = read_input(input_path)

        if not chapters:
            status_manager.update_status(thread_id, file_display, 0, "无数据")
            return None, [], []

        # 为当前文件生成输出路径
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)

        # 从输入路径提取文件名（不含扩展名）作为paper_id
        input_basename = os.path.basename(input_path)
        paper_id = os.path.splitext(input_basename)[0]

        # 生成输出文件名
        output_filename = f"{paper_id}.jsonl"
        output_path = os.path.join(output_dir, output_filename)

        # 检查输出文件是否已存在且不为空，如果存在则跳过处理
        if os.path.exists(output_path):
            existing_qa_count = count_qas_in_file(output_path)
            if existing_qa_count > 0:
                logger.info(f"输出文件已存在且包含 {existing_qa_count} 条QA，跳过处理: {output_path}")
                print(f"\n\033[1;33;40m【跳过】\033[0m 文件 \033[1;36;40m{file_display}\033[0m 的输出已存在（{existing_qa_count} 条QA），跳过处理")
                status_manager.update_status(thread_id, file_display, existing_qa_count, "已存在，跳过")
                # 读取已存在的QA数据用于统计
                existing_qas = []
                try:
                    with open(output_path, 'r', encoding='utf-8') as f:
                        for line in f:
                            if line.strip():
                                existing_qas.append(json.loads(line))
                except Exception:
                    pass
                # 返回空统计信息
                return output_path, existing_qas, [{"book_id": paper_id, "chunks_count": 0, "generated_total": existing_qa_count}]

        # 注意：如果文件已存在但为空，继续处理（支持断点续接功能）
        # process_book 函数会读取已处理的章节并跳过

        # 更新状态：正在处理（从文件读取当前QA数量）
        current_qa_count = count_qas_in_file(output_path)
        status_manager.update_status(thread_id, file_display, current_qa_count, "处理中")

        # 处理所有章节
        result = process_book(
            chapters,
            paper_id=paper_id,
            output_path=output_path,
            enable_diversity=args.enable_diversity_filter,
            simhash_dedup_hamming=args.simhash_dedup_hamming,
            enable_quality_filter=args.enable_quality_filter,
            min_quality_score=args.min_quality_score,
            max_curriculum_stage=args.max_curriculum_stage,
            thread_id=thread_id,
            file_display=file_display
        )
        qas = result.get("qas", [])
        stats = result.get("stats", {})

        # 从文件读取最新的QA数量（因为process_book已经实时写入了）
        final_qa_count = count_qas_in_file(output_path)
        status_manager.update_status(thread_id, file_display, final_qa_count, "完成")
        return output_path, qas, [stats]

    except Exception as e:
        logger.error(f"线程{thread_id}处理文件 {input_path} 异常: {e}")
        status_manager.update_status(thread_id, file_display, 0, f"错误: {str(e)[:30]}")
        return None, [], []

def update_status_from_files():
    """从输出文件读取QA数量并更新状态（后台线程定期调用）"""
    status = status_manager.get_status()
    output_dir = 'output'

    for thread_id, info in status.items():
        if info['status'] in ['处理中', '开始处理']:
            # 从文件名提取输出文件路径
            file_name = info['file']
            # file_name格式：20260105/9787040487039.md
            # 提取文件名部分（不含目录）
            if '/' in file_name:
                base_name = file_name.split('/')[-1]
            else:
                base_name = file_name
            # 去掉扩展名，添加.jsonl
            if '.' in base_name:
                base_name = base_name.rsplit('.', 1)[0]
            output_filename = f"{base_name}.jsonl"
            output_path = os.path.join(output_dir, output_filename)

            # 读取文件中的QA数量
            qa_count = count_qas_in_file(output_path)
            # 更新状态（保持原有状态文本）
            status_manager.update_status(thread_id, file_name, qa_count, info['status'])

def display_status_thread():
    """状态显示线程，每分钟更新一次"""
    first_update = True

    while status_manager.running:
        # 在显示前，先从文件更新状态
        update_status_from_files()

        time.sleep(5)  # 每5秒更新一次（提高更新频率，让用户看到实时进度）
        if not status_manager.running:
            break

        status = status_manager.get_status()
        if not status:
            continue

        # 显示所有线程状态（每个线程一行）
        lines = []
        for thread_id in sorted(status.keys()):
            info = status[thread_id]
            file_name = info['file']
            qa_count = info['qa_count']
            status_text = info['status']
            lines.append(f"线程{thread_id:2d} {status_text} 文件{file_name}，目前已生成{qa_count}条QA")

        # 如果不是第一次更新，需要向上移动光标覆盖之前的内容
        if not first_update and lines:
            # 向上移动光标（回到之前显示的位置）
            sys.stdout.write('\033[{}A'.format(len(lines)))

        # 打印新的状态（覆盖旧内容）
        for line in lines:
            sys.stdout.write('\033[K' + line + '\n')  # \033[K清除到行尾

        # 不需要补充空行，因为预先打印的空行数量已经根据实际线程数/文件数设置好了
        # 如果预先打印的空行数多于实际线程数，多余的会被覆盖掉

        sys.stdout.flush()
        first_update = False

# --- 主处理函数 ---
def process_book(
    chapters: List[Dict[str, Any]],
    paper_id: str,
    output_path: str = None,
    enable_diversity: bool = True,
    simhash_dedup_hamming: int = 6,
    enable_quality_filter: bool = False,
    min_quality_score: float = 60.0,
    max_curriculum_stage: int = None,
    thread_id: int = None,
    file_display: str = None
) -> Dict[str, Any]:
    """处理章节列表，返回问答对和统计信息

    Args:
        chapters: 章节列表（从JSON文件读取，包含books_ID, chapter_title, context, length）
        paper_id: 输入文件名（不含扩展名）
        output_path: 输出文件路径（如果提供，将实时写入）
        enable_diversity: 是否启用多样性过滤
        simhash_dedup_hamming: SimHash去重阈值
        enable_quality_filter: 是否启用质量过滤
        min_quality_score: 最小质量分数阈值
        max_curriculum_stage: curriculum最大阶段（如果提供，将在实时写入时应用过滤）

    Returns:
        {
          "book_id": ...,
          "qas": [...],
          "stats": {...}
        }
    """
    # 高亮显示开始处理
    print("\n" + "=" * 80)
    print(f"\033[1;33;40m【开始处理】\033[0m 文件: \033[1;32;40m{paper_id}\033[0m，共 {len(chapters)} 个章节")
    print("=" * 80)
    logger.info(f"开始处理文件: {paper_id}，共 {len(chapters)} 个章节")

    generator = SFTQuestionGenerator()
    quality_scorer = QualityScorer()

    try:
        # 读取已处理的章节（断点续接功能）
        processed_chapters = get_processed_chapters(output_path) if output_path else set()
        total_chunks = len(chapters)
        skipped_count = 0
        qas = []  # 初始化qas列表，用于收集所有章节的QA

        for chunk_idx, chapter in enumerate(chapters, 1):
            # 高亮显示正在处理的章节
            chapter_title = chapter.get('chapter_title', 'N/A')
            context = chapter.get('context', '')
            length = chapter.get('length', 0)
            books_ID = chapter.get('books_ID', paper_id)

            # 跳过"目录"章节，不生成QA
            if chapter_title.strip() in ["目录", "Contents", "CONTENTS", "目  录", "目　录"]:
                print(f"\n\033[1;33;40m【跳过】\033[0m \033[1;36;40m{chapter_title}\033[0m（目录章节，跳过）")
                logger.info(f"[{chunk_idx}/{total_chunks}] 跳过目录章节: {chapter_title}")
                skipped_count += 1
                continue

            # 先检查是否已处理
            if chapter_title in processed_chapters:
                skipped_count += 1
                print(f"\n\033[1;33;40m【跳过】\033[0m \033[1;36;40m{chapter_title}\033[0m（已处理，跳过）")
                logger.info(f"[{chunk_idx}/{total_chunks}] 章节 '{chapter_title}' 已处理，跳过")
                continue

            # 根据length字段判断是否符合生成QA的规则
            max_q_per_chunk = calculate_qa_count_per_chunk(length)

            # 如果不符合生成QA的规则，直接跳过，不做任何显示和输出
            if max_q_per_chunk == 0:
                skipped_count += 1
                logger.debug(f"[{chunk_idx}/{total_chunks}] 章节 '{chapter_title}' length {length} < 100，不符合生成QA规则，跳过")
                continue

            # 符合生成QA规则，显示正在处理（高亮显示）
            print(f"\n\033[1;35;40m正在处理章节\033[0m \033[1;36;40m\"{chapter_title}\"\033[0m (length: {length})")
            logger.info(f"[{chunk_idx}/{total_chunks}] 处理章节: {chapter_title}，length: {length}")

            try:
                # 构建chunk字典，用于生成QA
                chunk = {
                    'chunk_id': f"chunk_{chunk_idx:03d}",
                    'chunk_title': chapter_title,
                    'text': context
                }

                # 临时修改MAX_Q_PER_CHUNK以使用计算出的值
                global MAX_Q_PER_CHUNK
                original_max_q = MAX_Q_PER_CHUNK
                MAX_Q_PER_CHUNK = max_q_per_chunk

                # 生成QA（始终使用推理链生成）
                qs = generator.generate_for_chunk_with_reasoning(
                    chunk,
                    books_ID,
                    enable_diversity,
                    simhash_dedup_hamming,
                    enable_quality_filter,
                    min_quality_score
                )

                # 恢复原始MAX_Q_PER_CHUNK
                MAX_Q_PER_CHUNK = original_max_q

                # 为每个QA添加paper_id和chapter_title
                for qa in qs:
                    qa['paper_id'] = paper_id
                    qa['chapter_title'] = chapter_title

                # 应用curriculum stage过滤（如果指定）
                if max_curriculum_stage is not None:
                    qs = [
                        qa for qa in qs
                        if int(qa.get("curriculum_stage", 3)) <= max_curriculum_stage
                    ]

                # 实时写入（写入过滤后的QA）
                if output_path and qs:
                    for qa in qs:
                        save_qa_single(qa, output_path)
                    current_call_count = get_llm_call_count()
                    # 显示输出的QA的chapter_title
                    if qs:
                        qa_chapter_title = qs[0].get('chapter_title', chapter_title)
                        print(f"输出的QA\"chapter_title\"：\"{qa_chapter_title}\"")
                    print(f"  \033[1;32;40m✓\033[0m 已生成并实时写入 \033[1;33;40m{len(qs)}\033[0m 个QA | "
                          f"\033[1;36;40m累计LLM调用: {current_call_count}次\033[0m")

                    # 实时更新状态（从文件读取QA数量）
                    if thread_id is not None and file_display and output_path and status_manager:
                        try:
                            current_qa_count = count_qas_in_file(output_path)
                            status_manager.update_status(thread_id, file_display, current_qa_count, "处理中")
                        except Exception:
                            pass  # 如果更新失败，不影响主流程

                qas.extend(qs)

            except Exception as e:
                print(f"  \033[1;31;40m✗\033[0m 章节处理失败: {e}")
                logger.error(f"Chunk '{chapter_title}' 异常: {e}")

        # 统计（原始生成）
        gen_total = len(qas)

        # 汇总 diversity 过滤统计（按 chunk 去重）
        seen_chunk_keys = set()
        div_before_sum, div_after_sum = 0, 0
        simhash_before_sum, simhash_after_sum = 0, 0
        quality_before_sum, quality_after_sum = 0, 0

        for qa in qas:
            k = (qa.get("chapter_title"), qa.get("generation_type"))
            if k in seen_chunk_keys:
                continue
            seen_chunk_keys.add(k)
            # 统计推理型生成（现在统一使用推理型生成）
            if qa.get("generation_type") == "推理型":
                if "diversity_before" in qa and "diversity_after" in qa:
                    div_before_sum += int(qa.get("diversity_before", 0))
                    div_after_sum += int(qa.get("diversity_after", 0))
                if "simhash_before" in qa and "simhash_after" in qa:
                    simhash_before_sum += int(qa.get("simhash_before", 0))
                    simhash_after_sum += int(qa.get("simhash_after", 0))
                if "quality_before" in qa and "quality_after" in qa:
                    quality_before_sum += int(qa.get("quality_before", 0))
                    quality_after_sum += int(qa.get("quality_after", 0))

        # 图书级质量过滤（仅在chunk级别未过滤时执行，避免重复过滤）
        # 注意：如果chunk级别已经过滤，这里不再重复过滤，只做统计
        qas_after_quality = qas
        book_quality_before = len(qas_after_quality)
        # 检查是否在chunk级别已经过滤过（通过检查是否有quality_before/quality_after字段）
        chunk_level_filtered = any(
            "quality_before" in qa and "quality_after" in qa
            for qa in qas[:min(5, len(qas))]  # 只检查前几个样本
        ) if qas else False

        if enable_quality_filter and not chunk_level_filtered:
            # 只在chunk级别未过滤时才在book级别过滤
            qas_after_quality = quality_scorer.filter_by_quality(qas_after_quality, min_quality_score)
        book_quality_after = len(qas_after_quality)

        # 获取当前LLM调用次数
        current_llm_calls = get_llm_call_count()

        # 图书级统计对象
        stats = {
            "chunks_count": len(chapters),
            "skipped_chunks": skipped_count,  # 跳过的章节数量（断点续接）
            "generated_total": int(gen_total),
            "book_quality_before": int(book_quality_before),
            "book_quality_after": int(book_quality_after),
            "diversity_before_sum": int(div_before_sum),
            "diversity_after_sum": int(div_after_sum),
            "simhash_before_sum": int(simhash_before_sum),
            "simhash_after_sum": int(simhash_after_sum),
            "quality_before_sum": int(quality_before_sum),
            "quality_after_sum": int(quality_after_sum),
            "llm_call_count": int(current_llm_calls),
        }

        # 质量统计
        quality_stats = quality_scorer.get_quality_statistics(qas_after_quality)
        stats.update(quality_stats)

        logger.info(
            f"[{paper_id}] 统计: chunks={len(chapters)}, generated={gen_total}, "
            f"quality={book_quality_before}->{book_quality_after}, "
            f"diversity(sum)={div_before_sum}->{div_after_sum}, "
            f"simhash(sum)={simhash_before_sum}->{simhash_after_sum}, "
            f"avg_score={quality_stats.get('average_score', 0):.1f}, "
            f"LLM调用={current_llm_calls}次"
        )

        # 将book_id添加到stats中，方便后续统计
        stats["book_id"] = paper_id
        return {"book_id": paper_id, "qas": qas_after_quality, "stats": stats}

    except Exception as e:
        logger.error(f"处理图书 '{book_id}' 时发生错误: {e}", exc_info=True)
        return {"book_id": book_id, "qas": [], "stats": {}}

# --- 采样策略 ---
# def sample_qas_with_strategy(
#     qas: List[Dict[str, Any]],
#     max_q: int,
#     difficulty_target: Optional[Dict[str, int]] = None,
# ) -> List[Dict[str, Any]]:
#     """按难度配比采样问答对"""
#     if not qas:
#         return []

#     if difficulty_target is None:
#         difficulty_target = {"easy": 1, "medium": 3, "hard": 1}

#     buckets = {"easy": [], "medium": [], "hard": []}
#     for qa in qas:
#         d = qa.get("difficulty", "medium")
#         if d not in buckets:
#             d = "medium"
#         buckets[d].append(qa)

#     total_target = sum(difficulty_target.values())
#     if total_target == 0:
#         total_target = 1
#     scale = max_q / total_target

#     scaled_target = {}
#     for d, v in difficulty_target.items():
#         scaled_target[d] = max(0, int(round(v * scale)))

#     current_sum = sum(scaled_target.values())
#     if current_sum > max_q:
#         for d in ["hard", "medium", "easy"]:
#             while current_sum > max_q and scaled_target[d] > 0:
#                 scaled_target[d] -= 1
#                 current_sum -= 1
#     elif current_sum < max_q:
#         for d in ["medium", "easy", "hard"]:
#             while current_sum < max_q:
#                 scaled_target[d] += 1
#                 current_sum += 1
#                 if current_sum == max_q:
#                     break

#     selected = []

#     def pop_some(bucket_list: List[Dict[str, Any]], need: int) -> List[Dict[str, Any]]:
#         picked = []
#         for qa in bucket_list:
#             if len(picked) >= need:
#                 break
#             picked.append(qa)
#         return picked

#     for d in ["easy", "medium", "hard"]:
#         need = scaled_target.get(d, 0)
#         cand = buckets[d]
#         picked = pop_some(cand, need)
#         selected.extend(picked)

#     if len(selected) < max_q:
#         remaining_qas = [qa for qa in qas if qa not in selected]
#         for qa in remaining_qas:
#             if len(selected) >= max_q:
#                 break
#             selected.append(qa)

#     return selected[:max_q]

def sample_qas_with_strategy(
    qas: List[Dict[str, Any]],
    max_q: int,
    difficulty_target: Optional[Dict[str, int]] = None,
    max_stage: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    按难度配比 + Reasoning-aware curriculum 采样问答对

    Args:
        qas: 输入问答对列表
        max_q: 目标采样数量
        difficulty_target: 难度配比
        max_stage: curriculum 最大阶段（None 表示不过滤）
    """
    if not qas:
        return []

    # =========================
    # NEW: curriculum stage 过滤
    # =========================
    if max_stage is not None:
        qas = [
            qa for qa in qas
            if int(qa.get("curriculum_stage", 3)) <= int(max_stage)
        ]
        if not qas:
            return []

    # -------------------------
    # 原有逻辑（几乎未改）
    # -------------------------
    if difficulty_target is None:
        difficulty_target = {"easy": 1, "medium": 3, "hard": 1}

    buckets = {"easy": [], "medium": [], "hard": []}
    for qa in qas:
        d = qa.get("difficulty", "medium")
        if d not in buckets:
            d = "medium"
        buckets[d].append(qa)

    total_target = sum(difficulty_target.values())
    if total_target == 0:
        total_target = 1

    scale = max_q / total_target

    scaled_target = {}
    for d, v in difficulty_target.items():
        scaled_target[d] = max(0, int(round(v * scale)))

    current_sum = sum(scaled_target.values())

    # 防止溢出
    if current_sum > max_q:
        for d in ["hard", "medium", "easy"]:
            while current_sum > max_q and scaled_target[d] > 0:
                scaled_target[d] -= 1
                current_sum -= 1
    elif current_sum < max_q:
        for d in ["medium", "easy", "hard"]:
            while current_sum < max_q:
                scaled_target[d] += 1
                current_sum += 1
                if current_sum == max_q:
                    break

    selected = []

    def pop_some(bucket_list: List[Dict[str, Any]], need: int) -> List[Dict[str, Any]]:
        picked = []
        for qa in bucket_list:
            if len(picked) >= need:
                break
            picked.append(qa)
        return picked

    for d in ["easy", "medium", "hard"]:
        need = scaled_target.get(d, 0)
        cand = buckets[d]
        picked = pop_some(cand, need)
        selected.extend(picked)

    # 不足补齐
    if len(selected) < max_q:
        remaining_qas = [qa for qa in qas if qa not in selected]
        for qa in remaining_qas:
            if len(selected) >= max_q:
                break
            selected.append(qa)

    return selected[:max_q]


# --- 主程序 ---
def main():
    parser = argparse.ArgumentParser(description='图书SFT问答对生成器')
    parser.add_argument(
        '--input',
        type=str,
        required=False,
        default='examples',
        help='输入路径：可为Excel文件（读取sheet "OCR" 的 D 列）、JSONL文件或包含多个JSON/JSONL的目录'
    )
    parser.add_argument(
        '--output',
        type=str,
        required=False,
        default=None,
        help='输出JSONL文件路径（如果不指定，将根据输入文件名自动生成到output/目录）'
    )
    parser.add_argument('--model', type=str, default=None, help='使用的模型')
    parser.add_argument('--max-q-per-chunk', type=int, default=None, help='每章节最大问答数')
    parser.add_argument('--target-ids', type=str, nargs='*', default=[], help='目标图书ID列表')
    parser.add_argument('--sample-strategy', action='store_true', help='使用采样策略')
    parser.add_argument(
        '--max-curriculum-stage',
        type=int,
        default=None,
        help='curriculum 最大阶段（1/2/3，None表示不限制）'
    )
    parser.add_argument(
        '--enable-quality-filter',
        action='store_true',
        help='启用质量过滤'
    )
    parser.add_argument(
        '--min-quality-score',
        type=float,
        default=60.0,
        help='最小质量分数阈值（默认60.0）'
    )
    parser.add_argument(
        '--enable-diversity-filter',
        action='store_true',
        help='启用多样性过滤'
    )
    parser.add_argument(
        '--simhash-dedup-hamming',
        type=int,
        default=6,
        help='SimHash去重阈值（默认6，越小越严格）'
    )

    args = parser.parse_args()

    # 始终使用推理链生成

    # 更新全局变量
    global DEFAULT_MODEL, MAX_Q_PER_CHUNK
    if args.model is not None:
        DEFAULT_MODEL = args.model
    if args.max_q_per_chunk is not None:
        MAX_Q_PER_CHUNK = args.max_q_per_chunk

    logger.info("=" * 60)
    logger.info("开始执行图书SFT问答对生成流程")
    logger.info("=" * 60)

    # 检查API密钥是否配置
    if not OPENAI_API_KEY:
        error_msg = "错误：未设置 OPENAI_API_KEY 环境变量！\n" \
                   "请设置环境变量：export OPENAI_API_KEY='your-api-key-here'\n" \
                   "或创建 .env 文件并添加：OPENAI_API_KEY=${OPENAI_API_KEY}"
        print(f"\n\033[1;31;40m{error_msg}\033[0m\n")
        logger.error(error_msg)
        return

    # 根据输入文件名自动生成输出路径（如果未指定）
    if args.output is None:
        # 输出目录
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)

        # 从输入路径提取文件名（不含扩展名）
        input_basename = os.path.basename(args.input)
        input_name_without_ext = os.path.splitext(input_basename)[0]

        # 生成输出文件名
        output_filename = f"{input_name_without_ext}.jsonl"
        args.output = os.path.join(output_dir, output_filename)
        logger.info(f"未指定输出路径，自动生成: {args.output}")

    logger.info(f"输入路径: {args.input}")
    logger.info(f"使用模型: {DEFAULT_MODEL}")
    logger.info(f"每章节最大问答数: {MAX_Q_PER_CHUNK}")
    logger.info("推理链生成: 始终开启")

    # 重置LLM调用计数器
    reset_llm_call_count()
    print(f"\n\033[1;36;40m【LLM调用统计已初始化】\033[0m 开始统计LLM调用次数\n")

    # 检查输入是否为Excel文件
    input_paths = []
    if args.input.endswith('.xlsx') or args.input.endswith('.xls'):
        # 从Excel读取路径列表
        logger.info(f"检测到Excel文件，从sheet 'OCR' 的 D 列读取路径")
        excel_paths = read_excel_paths(args.input, sheet_name='OCR', column_index=3)

        # 查找每个路径对应的文件
        for excel_path in excel_paths:
            if not excel_path or excel_path.strip() == '':
                continue
            found_path = find_file_in_directory(excel_path.strip())
            if found_path:
                input_paths.append(found_path)
            else:
                logger.warning(f"未找到文件: {excel_path}")

        logger.info(f"从Excel读取到 {len(excel_paths)} 个路径，找到 {len(input_paths)} 个有效文件")
    else:
        # 检查输入是文件还是目录
        if os.path.isdir(args.input):
            # 如果是目录，获取目录下所有json/jsonl文件
            logger.info(f"输入为目录，将获取其中的所有JSON/JSONL文件: {args.input}")
            input_paths = []
            for fname in sorted(os.listdir(args.input)):
                if not (fname.endswith('.json') or fname.endswith('.jsonl')):
                    continue
                fpath = os.path.join(args.input, fname)
                if os.path.isfile(fpath):
                    input_paths.append(fpath)
            logger.info(f"目录下找到 {len(input_paths)} 个JSON/JSONL文件")
        else:
            # 单个文件
            input_paths = [args.input]

    if not input_paths:
        logger.error("未找到任何有效的输入文件")
        return

    # 初始化全局状态管理器
    global status_manager
    max_workers = 10  # 实际使用的线程数（每个线程处理1个文件）
    status_manager = ThreadStatusManager(max_threads=max_workers)
    logger.info(f"开始多线程处理，共 {len(input_paths)} 个文件，使用 {max_workers} 个线程")
    print(f"\n\033[1;36;40m【多线程处理启动】\033[0m 共 {len(input_paths)} 个文件，{max_workers} 个线程")
    print("="*80)
    print("线程状态（每5秒更新一次）：")
    # 预先打印空行，用于后续更新（根据实际文件数量，最多不超过实际线程数）
    num_lines = min(len(input_paths), max_workers)
    for i in range(num_lines):
        print()
    sys.stdout.flush()

    # 启动状态显示线程
    display_thread = threading.Thread(target=display_status_thread, daemon=True)
    display_thread.start()

    # 使用线程池处理所有文件
    all_qas = []
    all_stats = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_path = {}
        for idx, input_path in enumerate(input_paths):
            thread_id = idx + 1
            future = executor.submit(process_single_file, input_path, thread_id, args)
            future_to_path[future] = (input_path, thread_id)

        # 等待所有任务完成并收集结果
        for future in as_completed(future_to_path):
            input_path, thread_id = future_to_path[future]
            try:
                output_path, qas, stats = future.result()
                if output_path:
                    all_qas.extend(qas)
                    all_stats.extend(stats)
                    logger.info(f"线程{thread_id}完成: {os.path.basename(input_path)} -> {len(qas)}条QA")
            except Exception as e:
                logger.error(f"线程{thread_id}处理文件 {input_path} 时发生异常: {e}")

    # 停止状态显示线程
    status_manager.stop()
    time.sleep(2)  # 等待显示线程结束，确保最后一次更新完成

    # 显示最终状态
    final_status = status_manager.get_status()
    if final_status:
        print("\n" + "="*80)
        print("最终线程状态：")
        for thread_id in sorted(final_status.keys()):
            info = final_status[thread_id]
            file_name = info['file']
            qa_count = info['qa_count']
            status_text = info['status']
            print(f"线程{thread_id:2d} {status_text} 文件{file_name}，共生成{qa_count}条QA")
        print("="*80 + "\n")

    print("所有文件处理完成\n")

    # 统计不同生成类型的问答对数量
    reasoning_count = sum(1 for qa in all_qas if qa.get('generation_type') == '推理型')
    simple_count = 0  # 已移除简单生成模式，现在统一使用推理型生成
    unknown_count = len(all_qas) - reasoning_count - simple_count

    # 汇总统计
    quality_scorer = QualityScorer()
    quality_stats = quality_scorer.get_quality_statistics(all_qas)

    # 汇总图书级统计
    total_chunks = sum(s.get("chunks_count", 0) for s in all_stats)
    total_generated = sum(s.get("generated_total", 0) for s in all_stats)
    total_books = len(all_stats)
    total_llm_calls = get_llm_call_count()  # 获取LLM调用总次数

    # 注意：每个文件已经保存到各自的输出文件，这里只做统计
    # 如果用户指定了统一输出路径，也可以选择保存汇总结果（可选）
    if args.output and len(input_paths) > 1:
        # 如果指定了输出路径且处理了多个文件，可以选择保存汇总
        logger.info(f"检测到多个文件，每个文件已保存到独立输出文件")
    else:
        logger.info("所有QA已实时写入到各自的输出文件")
    logger.info("=" * 70)
    logger.info("流程执行完成！")
    logger.info("=" * 70)
    logger.info(f"【图书级别统计】")
    logger.info(f"  - 处理图书数: {total_books} 本")
    logger.info(f"  - 总chunks数: {total_chunks} 个")
    logger.info(f"  - 总生成QA数: {total_generated} 个")
    logger.info(f"  - LLM调用总次数: \033[1;36;40m{total_llm_calls}\033[0m 次")
    print(f"\n\033[1;36;40m【LLM调用统计】\033[0m 总调用次数: \033[1;33;40m{total_llm_calls}\033[0m 次")
    logger.info(f"【QA级别统计】")
    logger.info(f"  - 最终QA数: {len(all_qas)} 个")
    logger.info(f"  - 推理型问答对: {reasoning_count} 个")
    logger.info(f"  - 简单型问答对: {simple_count} 个")
    if unknown_count > 0:
        logger.info(f"  - 未分类问答对: {unknown_count} 个")
    logger.info(f"【质量统计】")
    logger.info(f"  - 平均质量分数: {quality_stats.get('average_score', 0):.2f}")
    logger.info(f"  - 最高分: {quality_stats.get('max_score', 0):.2f}")
    logger.info(f"  - 最低分: {quality_stats.get('min_score', 0):.2f}")
    logger.info(f"  - 优秀(≥90): {quality_stats.get('excellent_count', 0)} 个")
    logger.info(f"  - 良好(75-90): {quality_stats.get('good_count', 0)} 个")
    logger.info(f"  - 可接受(60-75): {quality_stats.get('acceptable_count', 0)} 个")
    logger.info(f"  - 较差(<60): {quality_stats.get('poor_count', 0)} 个")
    logger.info(f"  - 通过率(≥60): {quality_stats.get('pass_rate', 0):.2f}%")
    logger.info(f"输出文件: {args.output}")
    logger.info("=" * 70)

    # 保存统计信息到文件
    try:
        # 生成统计文件路径
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)

        # 从输入路径提取文件名作为统计文件名
        if args.input:
            input_basename = os.path.basename(args.input)
            input_name_without_ext = os.path.splitext(input_basename)[0]
            stats_filename = f"统计_{input_name_without_ext}.txt"
        else:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            stats_filename = f"统计_{timestamp}.txt"

        stats_file_path = os.path.join(output_dir, stats_filename)

        # 写入统计信息
        with open(stats_file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("图书QA生成统计报告\n")
            f.write("=" * 80 + "\n\n")

            # 总体统计
            f.write("【图书级别统计】\n")
            f.write(f"  - 处理图书数: {total_books} 本\n")
            f.write(f"  - 总chunks数: {total_chunks} 个\n")
            f.write(f"  - 总生成QA数: {total_generated} 个\n")
            f.write(f"  - LLM调用总次数: {total_llm_calls} 次\n")
            f.write("\n")

            f.write("【LLM调用统计】\n")
            f.write(f"  总调用次数: {total_llm_calls} 次\n")
            f.write("\n")

            f.write("【QA级别统计】\n")
            f.write(f"  - 最终QA数: {len(all_qas)} 个\n")
            f.write(f"  - 推理型问答对: {reasoning_count} 个\n")
            f.write(f"  - 简单型问答对: {simple_count} 个\n")
            if unknown_count > 0:
                f.write(f"  - 未分类问答对: {unknown_count} 个\n")
            f.write("\n")

            f.write("【质量统计】\n")
            f.write(f"  - 平均质量分数: {quality_stats.get('average_score', 0):.2f}\n")
            f.write(f"  - 最高分: {quality_stats.get('max_score', 0):.2f}\n")
            f.write(f"  - 最低分: {quality_stats.get('min_score', 0):.2f}\n")
            f.write(f"  - 优秀(≥90): {quality_stats.get('excellent_count', 0)} 个\n")
            f.write(f"  - 良好(75-90): {quality_stats.get('good_count', 0)} 个\n")
            f.write(f"  - 可接受(60-75): {quality_stats.get('acceptable_count', 0)} 个\n")
            f.write(f"  - 较差(<60): {quality_stats.get('poor_count', 0)} 个\n")
            f.write(f"  - 通过率(≥60): {quality_stats.get('pass_rate', 0):.2f}%\n")
            f.write("\n")

            f.write("=" * 80 + "\n")
            f.write("每本书详细统计\n")
            f.write("=" * 80 + "\n\n")

            # 每本书的详细统计
            for idx, stats in enumerate(all_stats, 1):
                book_id = stats.get('book_id', f'未知图书_{idx}')
                f.write(f"【图书 {idx}: {book_id}】\n")
                f.write("\n")

                # LLM调用统计
                book_llm_calls = stats.get('llm_call_count', 0)
                f.write("  【LLM调用统计】\n")
                f.write(f"    总调用次数: {book_llm_calls} 次\n")
                f.write("\n")

                # QA级别统计
                book_qa_count = stats.get('generated_total', 0)
                f.write("  【QA级别统计】\n")
                f.write(f"    - 最终QA数: {book_qa_count} 个\n")
                f.write("\n")

                # 质量统计
                book_avg_score = stats.get('average_score', 0)
                book_pass_rate = stats.get('pass_rate', 0)
                book_excellent = stats.get('excellent_count', 0)
                book_good = stats.get('good_count', 0)
                book_acceptable = stats.get('acceptable_count', 0)
                book_poor = stats.get('poor_count', 0)

                f.write("  【质量统计】\n")
                f.write(f"    - 平均质量分数: {book_avg_score:.2f}\n")
                f.write(f"    - 优秀(≥90): {book_excellent} 个\n")
                f.write(f"    - 良好(75-90): {book_good} 个\n")
                f.write(f"    - 可接受(60-75): {book_acceptable} 个\n")
                f.write(f"    - 较差(<60): {book_poor} 个\n")
                f.write(f"    - 通过率(≥60): {book_pass_rate:.2f}%\n")
                f.write("\n")

                # 其他统计信息
                chunks_count = stats.get('chunks_count', 0)
                skipped_chunks = stats.get('skipped_chunks', 0)
                f.write("  【其他统计】\n")
                f.write(f"    - 章节数: {chunks_count} 个\n")
                f.write(f"    - 跳过章节数: {skipped_chunks} 个\n")
                f.write("\n")

                f.write("-" * 80 + "\n\n")

        logger.info(f"统计信息已保存到: {stats_file_path}")
        print(f"\n\033[1;32;40m【统计文件已生成】\033[0m {stats_file_path}\n")

    except Exception as e:
        logger.error(f"保存统计信息失败: {e}")
        logger.warning("统计信息未保存到文件，但已显示在终端")

if __name__ == "__main__":
    main()
